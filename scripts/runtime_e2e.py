#!/usr/bin/env python3
"""Heavy ScanOps runtime E2E using real listeners, Nmap, API roles, and optional Chromium.

The harness is intentionally outside the normal pytest suite. It creates an isolated
``SCANOPS_DATA_DIR``, starts a real ScanOps server, and exercises full, limited, and
selected-finding rescans against loopback TCP/UDP listeners. All server/listener
processes, ports, and temporary data are cleaned in ``finally``.
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import os
import re
import shutil
import signal
import socket
import sqlite3
import subprocess
import sys
import tempfile
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

from process_control import close_kill_job, popen_tracked


ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "backend"
FRONTEND_DIST = ROOT / "frontend" / "dist"
ENGINE = ROOT / "engine"
HOST = "127.0.0.1"
TERMINAL_SCAN_STATES = {"done", "failed", "canceled", "interrupted"}


class RuntimeE2EError(AssertionError):
    """A failed runtime contract, with a concise CI-readable message."""


def log(message: str) -> None:
    print(f"[{time.strftime('%H:%M:%S')}] {message}", flush=True)


def require(condition: bool, message: str) -> None:
    if not condition:
        raise RuntimeE2EError(message)


def _capture_cleanup_error(errors: list[str], label: str, action) -> None:
    try:
        action()
    except Exception as exc:
        errors.append(f"{label}: {exc}")


def _wait_tcp_ready(port: int, label: str) -> None:
    deadline = time.monotonic() + 5
    while time.monotonic() < deadline:
        try:
            with socket.create_connection((HOST, port), timeout=0.5) as sock:
                sock.sendall(b"HEAD / HTTP/1.0\r\nHost: localhost\r\n\r\n")
                if sock.recv(64).startswith(b"HTTP/"):
                    return
        except OSError:
            time.sleep(0.05)
    raise RuntimeE2EError(f"{label} TCP listener did not become ready on {HOST}:{port}")


def _wait_udp_ready(port: int, label: str) -> None:
    deadline = time.monotonic() + 5
    with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as sock:
        sock.settimeout(0.5)
        while time.monotonic() < deadline:
            try:
                sock.sendto(b"runtime-e2e-ready", (HOST, port))
                payload, _address = sock.recvfrom(4096)
                if payload.startswith(b"ScanOps runtime E2E"):
                    return
            except TimeoutError:
                continue
    raise RuntimeE2EError(f"{label} UDP listener did not become ready on {HOST}:{port}")


class _LabHttpHandler(BaseHTTPRequestHandler):
    server_version = "ScanOpsRuntimeE2E/1.0"
    sys_version = ""

    def _reply(self, include_body: bool = True) -> None:
        body = b"ScanOps runtime E2E listener\n"
        self.send_response(200)
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        if include_body:
            self.wfile.write(body)

    def do_GET(self) -> None:  # noqa: N802 - BaseHTTPRequestHandler API
        self._reply()

    def do_HEAD(self) -> None:  # noqa: N802 - BaseHTTPRequestHandler API
        self._reply(include_body=False)

    def do_POST(self) -> None:  # noqa: N802 - Nmap may probe with POST
        self._reply()

    def log_message(self, _format: str, *_args) -> None:
        return


class _ReusableHttpServer(ThreadingHTTPServer):
    allow_reuse_address = True
    daemon_threads = True


class TcpLabListener:
    def __init__(self, label: str) -> None:
        self.label = label
        self.port = 0
        self._server: _ReusableHttpServer | None = None
        self._thread: threading.Thread | None = None

    def start(self) -> None:
        require(self._server is None, f"{self.label} TCP listener is already running")
        server = _ReusableHttpServer((HOST, self.port), _LabHttpHandler)
        if not self.port:
            self.port = int(server.server_address[1])
        self._server = server
        self._thread = threading.Thread(
            target=server.serve_forever,
            name=f"runtime-e2e-{self.label}-tcp",
            daemon=True,
        )
        self._thread.start()
        _wait_tcp_ready(self.port, self.label)
        log(f"listener {self.label}: tcp://{HOST}:{self.port} pid={os.getpid()}")

    def stop(self) -> None:
        server, thread = self._server, self._thread
        self._server = None
        self._thread = None
        if server is None:
            return
        server.shutdown()
        server.server_close()
        if thread is not None:
            thread.join(timeout=5)


class UdpLabListener:
    def __init__(self, label: str, avoid_ports: set[int] | None = None) -> None:
        self.label = label
        self.port = 0
        self._avoid_ports = set(avoid_ports or set())
        self._socket: socket.socket | None = None
        self._thread: threading.Thread | None = None
        self._stop = threading.Event()

    def _bind(self) -> socket.socket:
        while True:
            sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            sock.bind((HOST, self.port))
            chosen = int(sock.getsockname()[1])
            if chosen not in self._avoid_ports:
                self.port = chosen
                return sock
            sock.close()
            self.port = 0

    def start(self) -> None:
        require(self._socket is None, f"{self.label} UDP listener is already running")
        self._stop.clear()
        sock = self._bind()
        self._socket = sock
        self._thread = threading.Thread(
            target=self._serve,
            name=f"runtime-e2e-{self.label}-udp",
            daemon=True,
        )
        self._thread.start()
        _wait_udp_ready(self.port, self.label)
        log(f"listener {self.label}: udp://{HOST}:{self.port} pid={os.getpid()}")

    def _serve(self) -> None:
        sock = self._socket
        if sock is None:
            return
        while not self._stop.is_set():
            try:
                _payload, address = sock.recvfrom(65535)
                if self._stop.is_set():
                    return
                sock.sendto(b"ScanOps runtime E2E UDP listener\n", address)
            except OSError:
                if not self._stop.is_set():
                    raise
                return

    def stop(self) -> None:
        sock, thread = self._socket, self._thread
        self._socket = None
        self._thread = None
        if sock is None:
            return
        self._stop.set()
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as wake:
                wake.sendto(b"", (HOST, self.port))
        except OSError:
            pass
        if thread is not None:
            thread.join(timeout=2)
        sock.close()


class ApiClient:
    def __init__(self, base_url: str) -> None:
        self.base_url = base_url.rstrip("/")

    def request(
        self,
        method: str,
        path: str,
        *,
        token: str | None = None,
        payload=None,
        expected: int | tuple[int, ...] = 200,
        timeout: float = 30,
    ):
        headers = {"Accept": "application/json"}
        data = None
        if token:
            headers["Authorization"] = f"Bearer {token}"
        if payload is not None:
            headers["Content-Type"] = "application/json"
            data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        request = urllib.request.Request(
            self.base_url + path,
            data=data,
            headers=headers,
            method=method,
        )
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                status = response.status
                content_type = response.headers.get("Content-Type", "")
                body = response.read()
        except urllib.error.HTTPError as exc:
            status = exc.code
            content_type = exc.headers.get("Content-Type", "")
            body = exc.read()
        allowed = (expected,) if isinstance(expected, int) else expected
        if status not in allowed:
            text = body.decode("utf-8", errors="replace")[:1000]
            raise RuntimeE2EError(
                f"{method} {path} returned HTTP {status}, expected {allowed}: {text}"
            )
        if "application/json" in content_type:
            return json.loads(body.decode("utf-8"))
        return body


def _free_tcp_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.bind((HOST, 0))
        return int(sock.getsockname()[1])


def _find_nmap(explicit: str) -> str:
    candidates = [explicit] if explicit else []
    if found := shutil.which("nmap"):
        candidates.append(found)
    if os.name == "nt":
        candidates.extend([
            r"C:\Program Files (x86)\Nmap\nmap.exe",
            r"C:\Program Files\Nmap\nmap.exe",
        ])
    for candidate in candidates:
        if not candidate:
            continue
        path = Path(candidate).expanduser().resolve()
        if path.is_file():
            result = subprocess.run(
                [str(path), "--version"],
                capture_output=True,
                text=True,
                timeout=15,
                check=False,
            )
            if result.returncode == 0:
                return str(path)
    raise RuntimeE2EError("nmap was not found; install it or pass --nmap PATH")


def _check_sudo_for_nmap() -> None:
    if os.name == "nt" or not hasattr(os, "geteuid") or os.geteuid() == 0:
        return
    sudo = shutil.which("sudo")
    require(bool(sudo), "real SYN/UDP scans require root or passwordless sudo")
    result = subprocess.run(
        [sudo, "-n", "true"],
        capture_output=True,
        timeout=10,
        check=False,
    )
    require(result.returncode == 0, "passwordless sudo is required for real SYN/UDP scans")


def _tail(path: Path, limit: int = 6000) -> str:
    try:
        return path.read_text(encoding="utf-8", errors="replace")[-limit:]
    except OSError:
        return ""


def _start_server(
    data_dir: Path,
    api_port: int,
    nmap_path: str,
    log_path: Path,
) -> tuple[subprocess.Popen, object]:
    env = dict(os.environ)
    env.update({
        "PYTHONUTF8": "1",
        "PYTHONUNBUFFERED": "1",
        "SCANOPS_DATA_DIR": str(data_dir),
        "SCANOPS_HOST": HOST,
        "SCANOPS_PORT": str(api_port),
        "SCANOPS_NMAP_PATH": nmap_path,
        "SCANOPS_SCAN_SCOPE": f"{HOST}/32",
        "SCANOPS_ENGINE_DIR": str(ENGINE),
        "SCANOPS_FRONTEND_DIST": str(FRONTEND_DIST),
    })
    log_handle = log_path.open("wb")
    process = popen_tracked(
        [
            sys.executable,
            "-m",
            "uvicorn",
            "scanops.main:app",
            "--host",
            HOST,
            "--port",
            str(api_port),
            "--log-level",
            "info",
        ],
        cwd=BACKEND,
        env=env,
        stdout=log_handle,
        stderr=subprocess.STDOUT,
    )
    log(f"ScanOps server pid={process.pid} url=http://{HOST}:{api_port}")
    return process, log_handle


def _stop_process_tree(process: subprocess.Popen | None) -> None:
    if process is None:
        return
    job_close_error: Exception | None = None
    job_closed = False
    if os.name == "nt":
        try:
            job_closed = close_kill_job(process)
        except Exception as exc:
            job_close_error = exc
        descendants = [] if job_closed else _windows_descendant_pids(process.pid)
        if not job_closed:
            if process.poll() is None:
                subprocess.run(
                    ["taskkill", "/PID", str(process.pid), "/T", "/F"],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                    check=False,
                    timeout=20,
                )
            for pid in reversed(descendants):
                if _pid_is_running(pid):
                    subprocess.run(
                        ["taskkill", "/PID", str(pid), "/T", "/F"],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        check=False,
                        timeout=20,
                    )
    else:
        try:
            os.killpg(process.pid, signal.SIGTERM)
        except ProcessLookupError:
            pass
        try:
            process.wait(timeout=10)
        except subprocess.TimeoutExpired:
            try:
                os.killpg(process.pid, signal.SIGKILL)
            except ProcessLookupError:
                pass
    try:
        process.wait(timeout=5)
    except subprocess.TimeoutExpired:
        process.kill()
        process.wait(timeout=5)
    if os.name == "nt" and not job_closed:
        deadline = time.monotonic() + 5
        remaining = [pid for pid in descendants if _pid_is_running(pid)]
        while remaining and time.monotonic() < deadline:
            time.sleep(0.05)
            remaining = [pid for pid in remaining if _pid_is_running(pid)]
        if remaining:
            raise RuntimeE2EError(f"descendant processes remain after cleanup: {remaining}")
    if job_close_error is not None:
        raise RuntimeE2EError(f"kill-on-close job cleanup failed: {job_close_error}")


def _pid_is_running(pid: int) -> bool:
    if os.name == "nt":
        import ctypes
        from ctypes import wintypes

        kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)
        open_process = kernel32.OpenProcess
        open_process.argtypes = (wintypes.DWORD, wintypes.BOOL, wintypes.DWORD)
        open_process.restype = wintypes.HANDLE
        get_exit_code = kernel32.GetExitCodeProcess
        get_exit_code.argtypes = (wintypes.HANDLE, ctypes.POINTER(wintypes.DWORD))
        get_exit_code.restype = wintypes.BOOL
        close_handle = kernel32.CloseHandle
        close_handle.argtypes = (wintypes.HANDLE,)
        close_handle.restype = wintypes.BOOL

        handle = open_process(0x1000, False, pid)  # PROCESS_QUERY_LIMITED_INFORMATION
        if not handle:
            return ctypes.get_last_error() == 5  # ERROR_ACCESS_DENIED means it exists.
        try:
            exit_code = wintypes.DWORD()
            if not get_exit_code(handle, ctypes.byref(exit_code)):
                return True
            return exit_code.value == 259  # STILL_ACTIVE
        finally:
            close_handle(handle)
    try:
        os.kill(pid, 0)
        return True
    except PermissionError:
        return True
    except OSError:
        return False


def _windows_descendant_pids(root_pid: int) -> list[int]:
    powershell = shutil.which("powershell.exe") or shutil.which("powershell")
    require(bool(powershell), "PowerShell is required to verify Windows process cleanup")
    command = (
        "Get-CimInstance Win32_Process | ForEach-Object { "
        "'{0},{1}' -f $_.ProcessId,$_.ParentProcessId }"
    )
    result = subprocess.run(
        [powershell, "-NoProfile", "-Command", command],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
        check=False, timeout=15,
    )
    require(result.returncode == 0,
            f"Windows process snapshot failed: {result.stderr.strip()}")
    children: dict[int, list[int]] = {}
    for line in result.stdout.splitlines():
        try:
            pid_text, parent_text = line.strip().split(",", 1)
            pid, parent = int(pid_text), int(parent_text)
        except ValueError:
            continue
        children.setdefault(parent, []).append(pid)
    descendants: list[int] = []
    pending = list(children.get(root_pid, []))
    while pending:
        pid = pending.pop()
        if pid in descendants:
            continue
        descendants.append(pid)
        pending.extend(children.get(pid, []))
    return descendants


def _wait_for_health(api: ApiClient, process: subprocess.Popen, log_path: Path) -> None:
    deadline = time.monotonic() + 45
    last_error = ""
    while time.monotonic() < deadline:
        if process.poll() is not None:
            raise RuntimeE2EError(
                f"ScanOps exited during startup with {process.returncode}:\n{_tail(log_path)}"
            )
        try:
            health = api.request("GET", "/api/health", timeout=2)
            require(health.get("ready") is True, f"health not ready: {health}")
            return
        except (OSError, RuntimeE2EError, urllib.error.URLError) as exc:
            last_error = str(exc)
            time.sleep(0.25)
    raise RuntimeE2EError(f"ScanOps health timeout: {last_error}\n{_tail(log_path)}")


def _initial_admin_password(data_dir: Path) -> str:
    credential_file = data_dir / "INITIAL_ADMIN.txt"
    deadline = time.monotonic() + 10
    while time.monotonic() < deadline and not credential_file.exists():
        time.sleep(0.1)
    require(credential_file.is_file(), "INITIAL_ADMIN.txt was not created")
    text = credential_file.read_text(encoding="utf-8")
    match = re.search(r"비밀번호:\s*(\S+)", text)
    require(match is not None, "INITIAL_ADMIN.txt password line could not be parsed")
    return match.group(1)


def _login(api: ApiClient, username: str, password: str) -> str:
    response = api.request(
        "POST",
        "/api/auth/login",
        payload={"username": username, "password": password},
    )
    token = response.get("token")
    require(bool(token), f"login did not return a token for {username}")
    return token


def _wait_scan(
    api: ApiClient,
    token: str,
    scan_id: int,
    timeout: int,
    server_log: Path,
) -> dict:
    deadline = time.monotonic() + timeout
    scan = None
    while time.monotonic() < deadline:
        scan = api.request("GET", f"/api/scans/{scan_id}", token=token)
        if scan["status"] in TERMINAL_SCAN_STATES:
            break
        time.sleep(0.5)
    require(scan is not None and scan["status"] in TERMINAL_SCAN_STATES,
            f"scan #{scan_id} did not finish within {timeout}s")
    stages = api.request("GET", f"/api/scans/{scan_id}/stages", token=token)
    require(
        scan["status"] == "done",
        f"scan #{scan_id} ended as {scan['status']} "
        f"({stages.get('failure_code')}: {stages.get('failure_message')}):\n{_tail(server_log)}",
    )
    require(stages.get("kind") == "staged", f"scan #{scan_id} was not staged")
    require(stages.get("timeline_available") is True, f"scan #{scan_id} has no stage timeline")
    return stages


def _run_staged(
    api: ApiClient,
    token: str,
    *,
    name: str,
    ports: str,
    include_udp: bool,
    timeout: int,
    server_log: Path,
) -> tuple[int, dict]:
    options = ["fast", "version_light"]
    if include_udp:
        options.append("udp")
    started = api.request(
        "POST",
        "/api/scans/run-staged",
        token=token,
        payload={
            "name": name,
            "targets": [HOST],
            "options": options,
            "ports": ports,
            "nse": ["banner", "http-server-header"],
            "batch_size": 1,
            "discovery": "pn",
        },
    )
    scan_id = int(started["id"])
    log(f"scan #{scan_id} {name}: {ports}")
    return scan_id, _wait_scan(api, token, scan_id, timeout, server_log)


def _run_selected_rescan(
    api: ApiClient,
    token: str,
    finding_id: int,
    timeout: int,
    server_log: Path,
) -> tuple[int, dict]:
    started = api.request(
        "POST",
        "/api/findings/rescan",
        token=token,
        payload={"finding_ids": [finding_id], "options": ["version_light"]},
    )
    scan_id = int(started["scan_id"])
    log(f"scan #{scan_id} selected finding #{finding_id}")
    return scan_id, _wait_scan(api, token, scan_id, timeout, server_log)


def _findings(api: ApiClient, token: str) -> list[dict]:
    return api.request("GET", f"/api/findings?host={HOST}&state=", token=token)


def _finding(api: ApiClient, token: str, port: int, proto: str) -> dict:
    matches = [
        row for row in _findings(api, token)
        if int(row["port"]) == port and row["proto"] == proto
    ]
    require(len(matches) == 1, f"expected one finding for {HOST}:{port}/{proto}, got {len(matches)}")
    return matches[0]


def _assert_active(finding: dict, *, reopened: bool | None = None) -> None:
    require(finding["state"] in ("open", "open|filtered"),
            f"finding {finding['finding_key']} is not active: {finding['state']}")
    if reopened is not None:
        require(bool(finding["reopened"]) is reopened,
                f"finding {finding['finding_key']} reopened={finding['reopened']}, expected {reopened}")


def _stable_signature(finding: dict) -> dict:
    return {
        key: finding[key]
        for key in (
            "state", "reopened", "status", "service", "product", "version", "server",
            "display_identity", "banner", "cpe", "fingerprint", "identification", "last_seen",
        )
    }


def _check_server_consumers(
    api: ApiClient,
    viewer_token: str,
    auditor_token: str,
    finding: dict,
) -> dict:
    """Prove one real Server observation reaches every user-facing consumer."""
    finding_id = int(finding["id"])
    port = int(finding["port"])
    server = str(finding.get("server") or "")
    service = str(finding.get("service") or "")
    require("scanopsruntimee2e" in server.lower(),
            f"real HTTP Server header was not captured: {server!r}")
    require(finding.get("display_identity") == server,
            f"display identity did not prefer Server: {finding}")

    search_query = urllib.parse.urlencode({"q": server, "state": ""})
    search_rows = api.request(
        "GET", f"/api/findings?{search_query}", token=viewer_token,
    )
    require(any(int(row["id"]) == finding_id for row in search_rows),
            f"Server search did not return finding #{finding_id}: {search_rows}")

    dept = "Runtime Server Team"
    api.request(
        "PATCH", f"/api/findings/{finding_id}", token=auditor_token,
        payload={"dept": dept},
    )
    preview_query = urllib.parse.urlencode({"dept": dept})
    preview = api.request(
        "GET", f"/api/notifications/preview?{preview_query}", token=viewer_token,
    )
    require(server in preview["body"],
            f"notification preview omitted Server identity: {preview['body']!r}")
    if service and service != server:
        require(f"(서비스: {service})" in preview["body"],
                f"notification preview omitted taxonomy service context: {preview['body']!r}")

    feed = api.request("GET", "/api/events?limit=500", token=viewer_token)
    event = next(
        (item for item in feed["items"] if int(item["finding_id"]) == finding_id),
        None,
    )
    require(event is not None, f"global event feed omitted finding #{finding_id}")
    require(event.get("display_identity") == server and event.get("server") == server,
            f"global event feed did not prefer Server identity: {event}")
    require(event.get("service") == service,
            f"global event feed lost taxonomy service context: {event}")

    export_query = urllib.parse.urlencode({
        "cols": "port,display_identity,server,service", "q": server, "state": "",
    })
    csv_bytes = api.request(
        "GET", f"/api/findings/export?{export_query}&fmt=csv", token=viewer_token,
    )
    csv_rows = list(csv.DictReader(io.StringIO(csv_bytes.decode("utf-8-sig"))))
    csv_row = next((row for row in csv_rows if row["포트"] == str(port)), None)
    require(csv_row is not None, f"CSV export omitted Server finding on port {port}")
    require(csv_row["표시 식별"] == server and csv_row["Server"] == server,
            f"CSV export did not preserve Server identity: {csv_row}")
    require(csv_row["서비스"] == service,
            f"CSV export lost taxonomy service context: {csv_row}")

    xlsx_bytes = api.request(
        "GET", f"/api/findings/export?{export_query}&fmt=xlsx", token=viewer_token,
    )
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
        try:
            rows = list(workbook.active.iter_rows(values_only=True))
        finally:
            workbook.close()
    except ImportError as exc:
        raise RuntimeE2EError("openpyxl is required for Server XLSX verification") from exc
    headers = {value: index for index, value in enumerate(rows[0])}
    xlsx_row = next(
        (row for row in rows[1:] if int(row[headers["포트"]]) == port),
        None,
    )
    require(xlsx_row is not None, f"XLSX export omitted Server finding on port {port}")
    require(
        xlsx_row[headers["표시 식별"]] == server
        and xlsx_row[headers["Server"]] == server,
        f"XLSX export did not preserve Server identity: {xlsx_row}",
    )
    require(xlsx_row[headers["서비스"]] == service,
            f"XLSX export lost taxonomy service context: {xlsx_row}")

    return {
        "server": server,
        "display_identity": server,
        "service": service,
        "search": True,
        "notification_preview": True,
        "event_feed": True,
        "csv": True,
        "xlsx": True,
    }


def _assert_out_of_scope_unchanged(
    api: ApiClient,
    token: str,
    snapshots: dict[tuple[int, str], tuple[dict, int]],
) -> None:
    for (port, proto), (signature, event_count) in snapshots.items():
        current = _finding(api, token, port, proto)
        require(
            _stable_signature(current) == signature,
            f"out-of-scope finding changed for {port}/{proto}: "
            f"{_stable_signature(current)} != {signature}",
        )
        events = api.request("GET", f"/api/findings/{current['id']}/events", token=token)
        require(len(events) == event_count,
                f"out-of-scope finding gained events for {port}/{proto}")


def _check_exports(api: ApiClient, viewer_token: str, auditor_token: str, finding_id: int) -> dict:
    formula = "=RUNTIME_E2E()"
    api.request(
        "PATCH",
        f"/api/findings/{finding_id}",
        token=auditor_token,
        payload={"manual_note": formula},
    )
    query = urllib.parse.urlencode({"cols": "manual_note", "state": "open"})
    csv_bytes = api.request(
        "GET", f"/api/findings/export?{query}&fmt=csv", token=viewer_token,
    )
    require(csv_bytes.startswith(b"\xef\xbb\xbf"), "CSV export is missing UTF-8 BOM")
    csv_rows = list(csv.reader(io.StringIO(csv_bytes.decode("utf-8-sig"))))
    require(any("'=RUNTIME_E2E()" in row for row in csv_rows),
            "CSV formula value was not escaped with an apostrophe")

    xlsx_bytes = api.request(
        "GET", f"/api/findings/export?{query}&fmt=xlsx", token=viewer_token,
    )
    require(xlsx_bytes.startswith(b"PK\x03\x04"), "XLSX export is missing ZIP signature")
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=False)
        try:
            cells = [cell for row in workbook.active.iter_rows(values_only=False) for cell in row]
            require(any(cell.value == "'=RUNTIME_E2E()" and cell.data_type != "f" for cell in cells),
                    "XLSX formula value was not stored as safe text")
        finally:
            workbook.close()
    except ImportError as exc:
        raise RuntimeE2EError("openpyxl is required for XLSX runtime verification") from exc
    return {"csv_bom": True, "formula_escaped": True, "xlsx_zip": True}


def _seed_failed_stage_scan(data_dir: Path) -> dict:
    """Create one persisted terminal-stage fixture for browser history rendering."""
    stages = [
        {"stage": "discovery", "status": "done", "seconds": 0.1, "counts": {"live": 1}},
        {
            "stage": "tcp",
            "status": "error",
            "seconds": 0.1,
            "counts": {"errors": 1},
            "error": "runtime browser fixture stage failure",
        },
    ]
    now = time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime())
    db_path = data_dir / "scanops.db"
    require(db_path.is_file(), f"browser failure fixture database is missing: {db_path}")
    connection = sqlite3.connect(db_path, timeout=30)
    try:
        cursor = connection.execute(
            """
            INSERT INTO scan_runs (
                name, targets, command, status, started_at, finished_at,
                raw_xml_path, log_path, host_count, port_count, stages_json,
                failure_code, failure_message, created_by
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "runtime-browser-failed-stage",
                HOST,
                "단계스캔(엔진) · runtime browser fixture",
                "failed",
                now,
                now,
                "",
                "",
                1,
                0,
                json.dumps(stages, ensure_ascii=False),
                "engine_failed",
                "단계 스캔 중 오류가 발생했습니다.",
                None,
            ),
        )
        scan_id = int(cursor.lastrowid)
        connection.commit()
    finally:
        connection.close()
    scan_dir = data_dir / "scans" / f"scan_{scan_id}"
    scan_dir.mkdir(parents=True, exist_ok=True)
    (scan_dir / "spec.json").write_text(
        json.dumps({"fixture": "runtime-browser"}), encoding="utf-8",
    )
    return {"id": scan_id, "name": "runtime-browser-failed-stage"}


def _browser_go(page, label: str):
    menu = page.get_by_role("button", name="탐색 메뉴 열기", exact=True)
    mobile = menu.is_visible()
    if mobile:
        menu.click()
        sidebar = page.get_by_role("dialog", name="주 탐색", exact=True)
        sidebar.wait_for(state="visible", timeout=5_000)
    else:
        sidebar = page.get_by_role("complementary", name="주 탐색", exact=True)
        sidebar.wait_for(state="visible", timeout=5_000)
    navigation = sidebar.get_by_role("navigation")
    require(navigation.count() == 1, "primary navigation has no semantic navigation role")
    button = navigation.get_by_role(
        "button", name=re.compile(re.escape(label))
    )
    require(button.count() == 1, f"navigation button is ambiguous or missing: {label}")
    current_marker = page.locator("#primary-navigation nav.nav button").filter(has_text=label)
    require(current_marker.count() == 1, f"navigation DOM marker is ambiguous: {label}")
    button.click()
    page.locator(".topbar h2").wait_for(state="visible", timeout=10_000)
    require(current_marker.get_attribute("aria-current") == "page",
            f"navigation did not mark {label} as the current page")
    return button


def _browser_overflow_metrics(page, label: str) -> dict:
    metrics = page.evaluate(
        """
        () => {
          const measure = (node) => ({client: node.clientWidth, scroll: node.scrollWidth});
          return {
            document: measure(document.documentElement),
            body: measure(document.body),
            app: measure(document.querySelector('[data-app-shell]')),
            main: measure(document.querySelector('.main')),
          };
        }
        """
    )
    overflow = {
        key: value["scroll"] - value["client"]
        for key, value in metrics.items()
        if value and value["scroll"] > value["client"] + 1
    }
    require(not overflow, f"horizontal overflow at {label}: {overflow}")
    return metrics


def _browser_assert_dialog(page, dialog, opener, first_field, expect) -> None:
    expect(dialog).to_be_visible()
    require(dialog.get_attribute("aria-modal") == "true", "dialog is not aria-modal")
    shell = page.locator("[data-app-shell]")
    require(shell.evaluate("node => node.inert && node.hasAttribute('inert')"),
            "application shell is not inert while a password dialog is open")
    expect(first_field).to_be_focused()
    focusable = dialog.locator(
        'button:not([disabled]), input:not([disabled]), select:not([disabled]), '
        '[tabindex]:not([tabindex="-1"])'
    )
    require(focusable.count() >= 2, "password dialog has too few focusable controls")
    first_field.press("Shift+Tab")
    expect(focusable.last).to_be_focused()
    focusable.last.press("Tab")
    expect(first_field).to_be_focused()


def _browser_escape_dialog(page, dialog, opener, expect) -> None:
    page.keyboard.press("Escape")
    expect(dialog).to_be_hidden()
    expect(opener).to_be_focused()
    require(not page.locator("[data-app-shell]").evaluate("node => node.inert"),
            "application shell remained inert after closing a password dialog")


def _browser_scan_detail(page, fixture: dict, status: str, expect) -> None:
    table = page.locator("table.tbl")
    row = table.locator("tbody > tr").filter(has_text=fixture["name"]).first
    expect(row).to_be_visible()
    expect(row).to_contain_text(status)
    require(row.locator(".pill").count() >= 2,
            f"persisted stage timeline is missing for scan #{fixture['id']}")
    row.get_by_role("button", name="상세", exact=True).click()
    detail = page.locator(f"#scan-detail-{fixture['id']}")
    expect(detail).to_be_visible()
    expect(detail).to_contain_text(status)
    expect(detail).to_contain_text("단계 엔진")
    if status == "실패":
        expect(detail).to_contain_text("단계 스캔 중 오류가 발생했습니다.")
        expect(detail).to_contain_text("engine_failed")
        expect(detail).to_contain_text("runtime browser fixture stage failure")


def _browser_open_finding(page, finding_port: int, owner: str, expect):
    port_cell = page.locator("table.tbl").get_by_text(str(finding_port), exact=True).first
    expect(port_cell).to_be_visible()
    port_cell.click()
    drawer = page.locator(".drawer")
    expect(drawer).to_be_visible()
    expect(drawer).to_contain_text(owner)
    return drawer


def _browser_admin_checks(
    page,
    admin_password: str,
    finding_port: int,
    fixtures: dict,
    artifacts_dir: Path | None,
    expect,
) -> dict:
    password_opener = page.get_by_role("button", name="비밀번호 변경", exact=True)
    password_opener.click()
    self_dialog = page.get_by_role("dialog", name="비밀번호 변경", exact=True)
    current = self_dialog.get_by_placeholder("현재 비밀번호", exact=True)
    _browser_assert_dialog(page, self_dialog, password_opener, current, expect)
    current.fill(admin_password + "-wrong")
    next_password = "RuntimeBrowserNew-2026!"
    self_dialog.get_by_placeholder("새 비밀번호 (8자 이상)", exact=True).fill(next_password)
    self_dialog.get_by_placeholder("새 비밀번호 확인", exact=True).fill(next_password)
    self_dialog.get_by_role("button", name="변경", exact=True).click()
    error_alert = page.get_by_role("alert").filter(
        has_text="현재 비밀번호가 올바르지 않습니다."
    ).last
    expect(error_alert).to_be_visible()
    require(error_alert.get_attribute("aria-live") == "assertive",
            "error toast is not announced assertively")
    expect(self_dialog).to_be_visible()
    _browser_escape_dialog(page, self_dialog, password_opener, expect)

    _browser_go(page, "사용자")
    username = "runtime-browser-user"
    create_form = page.get_by_role("heading", name="새 사용자", exact=True).locator("..").locator("form")
    create_form.get_by_placeholder("아이디", exact=True).fill(username)
    create_form.get_by_placeholder("이름", exact=True).fill("Runtime Browser User")
    create_form.get_by_placeholder("비밀번호 (8자 이상)", exact=True).fill("RuntimeBrowser-2026!")
    create_form.locator("select").select_option("viewer")
    create_form.get_by_role("button", name="생성", exact=True).click()
    success_status = page.get_by_role("status").filter(has_text="사용자 생성됨").last
    expect(success_status).to_be_visible()
    require(success_status.get_attribute("aria-live") == "polite",
            "success toast is not announced politely")

    user_row = page.locator("table.tbl tbody tr").filter(has_text=username).first
    expect(user_row).to_be_visible()
    reset_opener = user_row.get_by_role("button", name="비밀번호 재설정", exact=True)
    reset_opener.click()
    reset_dialog = page.get_by_role("dialog", name=re.compile(re.escape(username)))
    reset_first = reset_dialog.get_by_placeholder("새 비밀번호 (8자 이상)", exact=True)
    _browser_assert_dialog(page, reset_dialog, reset_opener, reset_first, expect)
    _browser_escape_dialog(page, reset_dialog, reset_opener, expect)

    _browser_go(page, "스캔")
    _browser_scan_detail(page, fixtures["completed_scan"], "완료", expect)
    _browser_scan_detail(page, fixtures["failed_scan"], "실패", expect)

    xml_button = page.get_by_role("button", name="XML 가져오기(여러 개)", exact=True)
    xml_button.focus()
    with page.expect_file_chooser(timeout=5_000) as chooser_info:
        xml_button.press("Enter")
    chooser_info.value.set_files({
        "name": "runtime-empty.xml",
        "mimeType": "application/xml",
        "buffer": (
            b'<?xml version="1.0" encoding="UTF-8"?>'
            b'<nmaprun scanner="nmap" args="nmap -sn 127.0.0.1" start="0" '
            b'version="7.99" xmloutputversion="1.05">'
            b'<runstats><finished time="0" elapsed="0.00" summary="runtime"/>'
            b'</runstats></nmaprun>'
        ),
    })
    import_status = page.get_by_role("status").filter(has_text="가져옴").last
    expect(import_status).to_be_visible(timeout=10_000)
    expect(xml_button).to_be_focused(timeout=10_000)

    folder_button = page.get_by_role("button", name="폴더째 가져오기(XML+manifest)", exact=True)
    folder_button.focus()
    with page.expect_file_chooser(timeout=5_000) as chooser_info:
        folder_button.press("Space")
    chooser_info.value.element.dispatch_event("cancel", {"bubbles": True})
    expect(folder_button).to_be_focused()

    _browser_go(page, "이력")
    history_event = page.locator(".timeline .ev").filter(
        has_text=str(finding_port),
    ).filter(has_text=fixtures["server_identity"]).first
    expect(history_event).to_be_visible()
    expected_service = fixtures["server_service"]
    if expected_service and expected_service != fixtures["server_identity"]:
        expect(history_event).to_contain_text(f"(서비스: {expected_service})")

    _browser_go(page, "발견 관리")
    drawer = _browser_open_finding(
        page, finding_port, fixtures["asset"]["new_owner"], expect,
    )
    drawer.get_by_role("button", name="닫기", exact=True).click()
    expect(drawer).to_be_hidden()

    _browser_go(page, "자산대장")
    expect(page.get_by_text(fixtures["asset"]["old_owner"], exact=True)).to_be_visible()
    expect(page.get_by_text(fixtures["asset"]["new_owner"], exact=True)).to_be_visible()
    deletion = page.evaluate(
        """
        async (assetId) => {
          const token = localStorage.getItem('scanops_token');
          const response = await fetch(`/api/assets/${assetId}`, {
            method: 'DELETE', headers: {Authorization: `Bearer ${token}`},
          });
          return {status: response.status, body: await response.text()};
        }
        """,
        fixtures["asset"]["new_id"],
    )
    require(deletion["status"] == 204,
            f"browser asset deletion failed: {deletion}")

    _browser_go(page, "발견 관리")
    drawer = _browser_open_finding(
        page, finding_port, fixtures["asset"]["old_owner"], expect,
    )
    require(fixtures["asset"]["new_owner"] not in drawer.inner_text(),
            "finding retained deleted newest-asset attribution after re-entry")
    drawer.get_by_role("button", name="닫기", exact=True).click()
    _browser_go(page, "자산대장")
    expect(page.get_by_text(fixtures["asset"]["old_owner"], exact=True)).to_be_visible()
    require(page.get_by_text(fixtures["asset"]["new_owner"], exact=True).count() == 0,
            "deleted asset remained visible after re-entering the asset ledger")

    nav_labels = (
        "대시보드", "발견 관리", "히트맵", "규칙", "이력",
        "자산대장", "부서통보", "스캔", "사용자",
    )
    responsive = {}
    for width, height in ((390, 844), (768, 900), (1280, 800)):
        page.set_viewport_size({"width": width, "height": height})
        for label in nav_labels:
            _browser_go(page, label)
            _browser_overflow_metrics(page, f"{width}px/{label}")
        responsive[str(width)] = {"height": height, "views": list(nav_labels)}
        if artifacts_dir is not None:
            page.screenshot(
                path=str(artifacts_dir / f"browser-admin-{width}.png"), full_page=True,
            )
    return {
        "semantic_navigation": True,
        "account_actions": True,
        "toast_roles": {"success": "status/polite", "error": "alert/assertive"},
        "password_dialogs": ["self", "admin-reset"],
        "xml_focus": ["selection", "cancel"],
        "persisted_stage_details": [
            fixtures["completed_scan"]["id"], fixtures["failed_scan"]["id"],
        ],
        "asset_attribution_reentry": True,
        "history_server_identity": fixtures["server_identity"],
        "responsive": responsive,
    }


def _browser_checks(
    base_url: str,
    credentials: dict[str, tuple[str, str]],
    finding_port: int,
    fixtures: dict,
    artifacts_dir: Path | None,
) -> dict:
    try:
        from playwright.sync_api import expect, sync_playwright
    except ImportError as exc:
        raise RuntimeE2EError(
            "--browser requires scripts/requirements-runtime-e2e.txt and playwright install chromium"
        ) from exc

    checked = []
    admin_checks = {}
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        try:
            for role in ("admin", "auditor", "viewer"):
                username, password = credentials[role]
                context = browser.new_context(viewport={"width": 1280, "height": 800})
                page = context.new_page()
                try:
                    page.goto(base_url, wait_until="domcontentloaded", timeout=30_000)
                    page.get_by_placeholder("아이디").fill(username)
                    page.get_by_placeholder("••••••••").fill(password)
                    page.get_by_role("button", name="접속").click()
                    page.locator("[data-app-shell]").wait_for(state="visible", timeout=20_000)
                    who = page.locator(".who").inner_text()
                    require(role in who, f"browser role marker missing for {role}: {who!r}")

                    sidebar = page.get_by_role("complementary", name="주 탐색", exact=True)
                    require(sidebar.count() == 1, f"semantic sidebar is missing for {role}")
                    nav = sidebar.get_by_role("navigation")
                    require(nav.count() == 1, f"semantic navigation is missing for {role}")
                    require(page.get_by_role("button", name="비밀번호 변경", exact=True).count() == 1,
                            f"password account action is missing for {role}")
                    require(page.get_by_role("button", name="로그아웃", exact=True).count() == 1,
                            f"logout account action is missing for {role}")
                    user_nav_count = nav.get_by_role(
                        "button", name=re.compile(r"사용자$")
                    ).count()
                    require((user_nav_count == 1) is (role == "admin"),
                            f"browser user-management visibility is wrong for {role}")
                    nav.get_by_role("button", name=re.compile(r"스캔$")).click()
                    page.get_by_role("heading", name="스캔", exact=True).wait_for(timeout=10_000)
                    run_panel_count = page.get_by_role(
                        "heading", name="스캔 실행", exact=True
                    ).count()
                    require((run_panel_count == 1) is (role in ("admin", "auditor")),
                            f"browser scan controls visibility is wrong for {role}")

                    nav.get_by_role("button", name=re.compile("발견 관리")).click()
                    page.get_by_role("heading", name="발견 관리").wait_for(timeout=10_000)
                    page.locator("table.tbl").wait_for(state="visible", timeout=10_000)
                    try:
                        page.locator("table.tbl").get_by_text(
                            str(finding_port), exact=True,
                        ).first.wait_for(state="visible", timeout=10_000)
                    except Exception as exc:
                        raise RuntimeE2EError(
                            f"browser findings table does not expose port {finding_port} for {role}"
                        ) from exc
                    rescan_count = page.get_by_role("button", name=re.compile("선택 재스캔")).count()
                    require((rescan_count >= 1) is (role in ("admin", "auditor")),
                            f"browser rescan visibility is wrong for {role}")
                    if role == "admin":
                        admin_checks = _browser_admin_checks(
                            page, password, finding_port, fixtures, artifacts_dir, expect,
                        )
                    if artifacts_dir is not None:
                        page.screenshot(path=str(artifacts_dir / f"browser-{role}.png"), full_page=True)
                    checked.append(role)
                except Exception:
                    if artifacts_dir is not None:
                        page.screenshot(path=str(artifacts_dir / f"browser-{role}-failure.png"), full_page=True)
                    raise
                finally:
                    context.close()
        finally:
            browser.close()
    return {"checked_roles": checked, **admin_checks}


def _port_is_bindable(port: int, sock_type: int) -> bool:
    sock = socket.socket(socket.AF_INET, sock_type)
    try:
        if os.name == "nt" and hasattr(socket, "SO_EXCLUSIVEADDRUSE"):
            sock.setsockopt(socket.SOL_SOCKET, socket.SO_EXCLUSIVEADDRUSE, 1)
        elif sock_type == socket.SOCK_STREAM:
            sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        sock.bind((HOST, port))
        return True
    except OSError:
        return False
    finally:
        sock.close()


def _write_report(artifacts_dir: Path | None, report: dict) -> None:
    if artifacts_dir is None:
        return
    artifacts_dir.mkdir(parents=True, exist_ok=True)
    (artifacts_dir / "runtime-e2e-result.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def run(args: argparse.Namespace, report: dict) -> None:
    require(BACKEND.is_dir() and ENGINE.is_dir(), f"not a ScanOps checkout: {ROOT}")
    if args.browser:
        require((FRONTEND_DIST / "index.html").is_file(),
                "--browser requires `npm ci && npm run build` in frontend")
    nmap_path = _find_nmap(args.nmap)
    _check_sudo_for_nmap()
    report["nmap"] = nmap_path

    selected = TcpLabListener("selected")
    outside = TcpLabListener("outside")
    udp = UdpLabListener("udp")
    listeners = [selected, outside, udp]
    process: subprocess.Popen | None = None
    log_handle = None
    server_log: Path | None = None
    primary_error: BaseException | None = None
    cleanup_errors: list[str] = []
    temp = tempfile.TemporaryDirectory(prefix="scanops_runtime_e2e_")
    run_root = Path(temp.name).resolve()
    data_dir = run_root / "data"
    data_dir.mkdir()
    api_port = args.api_port
    artifacts_dir = Path(args.artifacts_dir).resolve() if args.artifacts_dir else None
    if artifacts_dir is not None:
        artifacts_dir.mkdir(parents=True, exist_ok=True)
    report.update({
        "runtime_root": str(run_root),
        "listener_pid": os.getpid(),
        "scans": [],
    })

    try:
        selected.start()
        outside.start()
        udp._avoid_ports = {selected.port, outside.port}
        udp.start()
        if not api_port:
            api_port = _free_tcp_port()
        require(api_port not in (selected.port, outside.port),
                f"API TCP port conflicts with a lab listener: {api_port}")
        report["server_port"] = api_port
        report["ports"] = {
            "selected_tcp": selected.port,
            "outside_tcp": outside.port,
            "udp": udp.port,
        }

        server_log = run_root / "server.log"
        process, log_handle = _start_server(data_dir, api_port, nmap_path, server_log)
        report["server_pid"] = process.pid
        api = ApiClient(f"http://{HOST}:{api_port}")
        _wait_for_health(api, process, server_log)

        admin_password = _initial_admin_password(data_dir)
        passwords = {
            "admin": admin_password,
            "auditor": "RuntimeAuditor-2026!",
            "viewer": "RuntimeViewer-2026!",
        }
        admin_token = _login(api, "admin", passwords["admin"])
        api.request("POST", "/api/users", token=admin_token, expected=201, payload={
            "username": "runtime-auditor", "password": passwords["auditor"],
            "role": "auditor", "display_name": "Runtime Auditor",
        })
        api.request("POST", "/api/users", token=admin_token, expected=201, payload={
            "username": "runtime-viewer", "password": passwords["viewer"],
            "role": "viewer", "display_name": "Runtime Viewer",
        })
        auditor_token = _login(api, "runtime-auditor", passwords["auditor"])
        viewer_token = _login(api, "runtime-viewer", passwords["viewer"])
        api.request("GET", "/api/users", token=admin_token)
        api.request("GET", "/api/users", token=auditor_token, expected=403)
        api.request("GET", "/api/users", token=viewer_token, expected=403)
        api.request("GET", "/api/scans/options", token=viewer_token)
        api.request(
            "POST", "/api/scans/run-staged", token=viewer_token, expected=403,
            payload={"targets": [HOST], "ports": f"T:{selected.port}", "discovery": "pn"},
        )
        report["rbac"] = {
            "admin_users": 200,
            "auditor_users": 403,
            "viewer_users": 403,
            "viewer_scan_start": 403,
            "viewer_scan_options": 200,
        }

        full_ports = f"T:{selected.port},{outside.port},U:{udp.port}"
        scan_id, stages = _run_staged(
            api, auditor_token, name="runtime-full-open", ports=full_ports,
            include_udp=True, timeout=args.scan_timeout, server_log=server_log,
        )
        report["scans"].append({"id": scan_id, "mode": "full", "transition": "open"})
        require(stages["host_count"] == 1, "full open scan did not persist one host")
        selected_finding = _finding(api, auditor_token, selected.port, "tcp")
        outside_finding = _finding(api, auditor_token, outside.port, "tcp")
        udp_finding = _finding(api, auditor_token, udp.port, "udp")
        _assert_active(selected_finding, reopened=False)
        _assert_active(outside_finding, reopened=False)
        _assert_active(udp_finding, reopened=False)
        report["server_consumers"] = _check_server_consumers(
            api, viewer_token, auditor_token, selected_finding,
        )
        api.request("GET", "/api/findings?state=open", token=viewer_token)
        api.request(
            "POST", "/api/findings/rescan-command", token=viewer_token,
            payload={"finding_ids": [selected_finding["id"]]},
        )
        api.request(
            "PATCH", f"/api/findings/{selected_finding['id']}", token=viewer_token,
            expected=403, payload={"manual_note": "viewer must not write"},
        )
        api.request(
            "POST", "/api/findings/rescan", token=viewer_token, expected=403,
            payload={"finding_ids": [selected_finding["id"]]},
        )
        report["rbac"].update({
            "viewer_findings": 200,
            "viewer_rescan_command": 200,
            "viewer_finding_patch": 403,
            "viewer_rescan_start": 403,
            "auditor_scan_start": 200,
        })

        for listener in listeners:
            listener.stop()
        scan_id, _ = _run_staged(
            api, auditor_token, name="runtime-full-closed", ports=full_ports,
            include_udp=True, timeout=args.scan_timeout, server_log=server_log,
        )
        report["scans"].append({"id": scan_id, "mode": "full", "transition": "closed"})
        for port, proto in ((selected.port, "tcp"), (outside.port, "tcp"), (udp.port, "udp")):
            finding = _finding(api, auditor_token, port, proto)
            require(finding["state"] == "closed", f"full scan did not close {port}/{proto}")

        for listener in listeners:
            listener.start()
        scan_id, _ = _run_staged(
            api, auditor_token, name="runtime-full-reopened", ports=full_ports,
            include_udp=True, timeout=args.scan_timeout, server_log=server_log,
        )
        report["scans"].append({"id": scan_id, "mode": "full", "transition": "reopened"})
        for port, proto in ((selected.port, "tcp"), (outside.port, "tcp"), (udp.port, "udp")):
            _assert_active(_finding(api, auditor_token, port, proto), reopened=True)

        out_of_scope: dict[tuple[int, str], tuple[dict, int]] = {}
        for port, proto in ((outside.port, "tcp"), (udp.port, "udp")):
            finding = _finding(api, auditor_token, port, proto)
            events = api.request("GET", f"/api/findings/{finding['id']}/events", token=auditor_token)
            out_of_scope[(port, proto)] = (_stable_signature(finding), len(events))

        selected.stop()
        scan_id, _ = _run_staged(
            api, auditor_token, name="runtime-limited-closed", ports=f"T:{selected.port}",
            include_udp=False, timeout=args.scan_timeout, server_log=server_log,
        )
        report["scans"].append({"id": scan_id, "mode": "limited", "transition": "closed"})
        require(_finding(api, auditor_token, selected.port, "tcp")["state"] == "closed",
                "limited scan did not close selected TCP finding")
        _assert_out_of_scope_unchanged(api, auditor_token, out_of_scope)

        selected.start()
        scan_id, _ = _run_staged(
            api, auditor_token, name="runtime-limited-reopened", ports=f"T:{selected.port}",
            include_udp=False, timeout=args.scan_timeout, server_log=server_log,
        )
        report["scans"].append({"id": scan_id, "mode": "limited", "transition": "reopened"})
        selected_finding = _finding(api, auditor_token, selected.port, "tcp")
        _assert_active(selected_finding, reopened=True)
        _assert_out_of_scope_unchanged(api, auditor_token, out_of_scope)

        selected.stop()
        scan_id, _ = _run_selected_rescan(
            api, auditor_token, selected_finding["id"], args.scan_timeout, server_log,
        )
        report["scans"].append({"id": scan_id, "mode": "selected", "transition": "closed"})
        require(_finding(api, auditor_token, selected.port, "tcp")["state"] == "closed",
                "selected rescan did not close selected TCP finding")
        _assert_out_of_scope_unchanged(api, auditor_token, out_of_scope)

        selected.start()
        scan_id, _ = _run_selected_rescan(
            api, auditor_token, selected_finding["id"], args.scan_timeout, server_log,
        )
        report["scans"].append({"id": scan_id, "mode": "selected", "transition": "reopened"})
        selected_finding = _finding(api, auditor_token, selected.port, "tcp")
        _assert_active(selected_finding, reopened=True)
        _assert_out_of_scope_unchanged(api, auditor_token, out_of_scope)

        events = api.request(
            "GET", f"/api/findings/{selected_finding['id']}/events", token=auditor_token,
        )
        event_types = [event["type"] for event in events]
        require(event_types.count("CLOSED") >= 3,
                f"full/limited/selected CLOSED events missing: {event_types}")
        require(event_types.count("REOPENED") >= 3,
                f"full/limited/selected REOPENED events missing: {event_types}")
        report["lifecycle"] = {
            "full": "open -> closed -> reopened",
            "limited": "closed -> reopened",
            "selected": "closed -> reopened",
            "out_of_scope_unchanged": True,
            "event_types": event_types,
        }

        report["exports"] = _check_exports(
            api, viewer_token, auditor_token, selected_finding["id"],
        )
        if args.browser:
            completed_scan = api.request(
                "GET", f"/api/scans/{scan_id}", token=auditor_token,
            )
            failed_scan = _seed_failed_stage_scan(data_dir)
            old_asset = api.request(
                "POST", "/api/assets", token=admin_token, expected=201,
                payload={
                    "ip": HOST,
                    "hostname": "runtime-old-asset",
                    "dept": "Runtime Older Department",
                    "owner": "Runtime Older Owner",
                    "contact": "old@example.invalid",
                    "asset_no": "RUNTIME-OLD",
                },
            )
            new_asset = api.request(
                "POST", "/api/assets", token=admin_token, expected=201,
                payload={
                    "ip": HOST,
                    "hostname": "runtime-new-asset",
                    "dept": "Runtime Newer Department",
                    "owner": "Runtime Newer Owner",
                    "contact": "new@example.invalid",
                    "asset_no": "RUNTIME-NEW",
                },
            )
            require(
                _finding(api, auditor_token, selected.port, "tcp")["owner"]
                == "Runtime Newer Owner",
                "newest browser asset fixture did not become finding attribution",
            )
            fixtures = {
                "completed_scan": {"id": scan_id, "name": completed_scan["name"]},
                "failed_scan": failed_scan,
                "server_identity": selected_finding["server"],
                "server_service": selected_finding["service"],
                "asset": {
                    "old_id": int(old_asset["id"]),
                    "new_id": int(new_asset["id"]),
                    "old_owner": "Runtime Older Owner",
                    "new_owner": "Runtime Newer Owner",
                },
            }
            report["browser"] = _browser_checks(
                f"http://{HOST}:{api_port}",
                {
                    "admin": ("admin", passwords["admin"]),
                    "auditor": ("runtime-auditor", passwords["auditor"]),
                    "viewer": ("runtime-viewer", passwords["viewer"]),
                },
                selected.port,
                fixtures,
                artifacts_dir,
            )
            attributed = _finding(api, auditor_token, selected.port, "tcp")
            require(attributed["owner"] == fixtures["asset"]["old_owner"],
                    "deleted newest asset did not restore older finding attribution")
            assets = api.request("GET", "/api/assets", token=admin_token)
            require(all(int(asset["id"]) != fixtures["asset"]["new_id"] for asset in assets),
                    "browser-deleted asset still exists through the API")
            report["browser"]["asset_attribution_api"] = {
                "deleted": fixtures["asset"]["new_id"],
                "fallback": fixtures["asset"]["old_id"],
                "owner": attributed["owner"],
            }
        else:
            report["browser"] = {"skipped": True, "reason": "--browser was not requested"}
        report["ok"] = True
        log("runtime E2E assertions passed")
    except BaseException as exc:
        primary_error = exc
        report["ok"] = False
        report["error"] = f"{type(exc).__name__}: {exc}"
    finally:
        _capture_cleanup_error(
            cleanup_errors, "server process tree stop", lambda: _stop_process_tree(process),
        )
        if log_handle is not None:
            _capture_cleanup_error(cleanup_errors, "server log close", log_handle.close)
        for listener in listeners:
            _capture_cleanup_error(
                cleanup_errors, f"{listener.label} stop", listener.stop,
            )

        ports = [api_port] if api_port else []
        if report.get("ports"):
            ports.extend([selected.port, outside.port])
            for port in ports:
                if not _port_is_bindable(port, socket.SOCK_STREAM):
                    cleanup_errors.append(f"TCP port still bound: {port}")
            if not _port_is_bindable(udp.port, socket.SOCK_DGRAM):
                cleanup_errors.append(f"UDP port still bound: {udp.port}")

        if artifacts_dir is not None and server_log is not None and server_log.exists():
            _capture_cleanup_error(
                cleanup_errors,
                "server log artifact copy",
                lambda: shutil.copy2(server_log, artifacts_dir / "server.log"),
            )
        if artifacts_dir is not None and primary_error is not None:
            scans_dir = data_dir / "scans"
            if scans_dir.exists():
                _capture_cleanup_error(
                    cleanup_errors,
                    "failed scan artifact copy",
                    lambda: shutil.copytree(
                        scans_dir, artifacts_dir / "failed-scans", dirs_exist_ok=True,
                    ),
                )
        _capture_cleanup_error(cleanup_errors, "temporary data cleanup", temp.cleanup)
        if run_root.exists():
            cleanup_errors.append(f"temporary runtime directory remains: {run_root}")
        report["cleanup"] = {
            "server_process_tree_stopped": process is None or process.poll() is not None,
            "ports_released": not any("port still bound" in error for error in cleanup_errors),
            "temporary_data_removed": not run_root.exists(),
            "errors": cleanup_errors,
        }
        _capture_cleanup_error(
            cleanup_errors, "runtime report write", lambda: _write_report(artifacts_dir, report),
        )

    if cleanup_errors:
        cleanup_message = "; ".join(cleanup_errors)
        if primary_error is not None:
            raise RuntimeE2EError(f"{primary_error}; cleanup failures: {cleanup_message}") from primary_error
        raise RuntimeE2EError(f"cleanup failures: {cleanup_message}")
    if primary_error is not None:
        raise primary_error


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=r"""Local PowerShell (run from the repository root):
  .\backend\.venv\Scripts\python.exe -m pip install -r scripts\requirements-runtime-e2e.txt
  .\backend\.venv\Scripts\python.exe -m playwright install chromium
  .\backend\.venv\Scripts\python.exe scripts\runtime_e2e.py --browser --artifacts-dir "$env:TEMP\scanops-runtime-e2e"
""",
    )
    parser.add_argument("--nmap", default="", help="explicit nmap executable path")
    parser.add_argument("--api-port", type=int, default=0, help="ScanOps port (0 chooses a free port)")
    parser.add_argument("--scan-timeout", type=int, default=240, help="seconds allowed per real scan")
    parser.add_argument("--browser", action="store_true", help="also validate role-based Chromium UI")
    parser.add_argument("--artifacts-dir", default="", help="optional persistent report/log/screenshot directory")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    report: dict = {"ok": False, "started_at": time.strftime("%Y-%m-%dT%H:%M:%S%z")}
    try:
        run(args, report)
    except Exception as exc:
        report.setdefault("error", f"{type(exc).__name__}: {exc}")
        report["ok"] = False
        _write_report(Path(args.artifacts_dir).resolve() if args.artifacts_dir else None, report)
        print(f"RUNTIME E2E FAILED: {exc}", file=sys.stderr, flush=True)
        return 1
    report["finished_at"] = time.strftime("%Y-%m-%dT%H:%M:%S%z")
    _write_report(Path(args.artifacts_dir).resolve() if args.artifacts_dir else None, report)
    print(json.dumps(report, ensure_ascii=False, indent=2), flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
