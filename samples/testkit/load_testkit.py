"""테스트킷 시더 — 스캔 XML + dirty 자산대장을 실행 중인 ScanOps 에 적재.

스캔 XML 은 실제 백엔드 경로(`POST /api/scans/import`)로 넣는다.
자산대장은 브라우저 고급 임포터(`frontend/src/lib/assetImport.js`)의 정제 로직
 — 병합셀 해제(forward-fill) · 헤더행 자동감지 · 컬럼 자동매핑 · blank 토큰 정리 ·
   결합셀 분리 — 을 그대로 파이썬으로 포팅해 `POST /api/assets/bulk` 로 넣는다.
(올인원 번들의 임베디드 파이썬에는 openpyxl 이 사전설치되어 있어 브라우저 없이 시딩 가능.)

전제: 서버가 8770 에서 가동 중, admin 비밀번호를 안다(data/INITIAL_ADMIN.txt).
사용:
  python load_testkit.py <admin_password>
  python load_testkit.py <admin_password> --base http://127.0.0.1:8770
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import urllib.error
import urllib.request
from pathlib import Path

HERE = Path(__file__).resolve().parent

SCANS = [
    ("scan_01_baseline.xml", "기준 스캔"),
    ("scan_02_rescan.xml", "재스캔(조치검증 diff)"),
    ("scan_03_discovery_only.xml", "발견만(노출 0)"),
    ("scan_04_broken.xml", "깨진 XML(오류 경로)"),
]

# ---------------------------------------------------------------------------
# frontend/src/lib/assetImport.js 포팅 — 정제 로직(동일 동작)
# ---------------------------------------------------------------------------
ASSET_ALIASES = {
    "ip": ["ip", "아이피", "ipaddress", "ipaddr", "ip주소"],
    "asset_no": ["자산", "자산번호", "자산코드", "관리번호", "코드", "번호", "asset", "assetid"],
    "dept": ["부서", "부서명", "관리부서", "소속", "조직", "팀"],
    "owner": ["담당", "담당자", "관리자", "책임", "관리담당", "관리책임", "소유자"],
    "contact": ["연락처", "전화", "전화번호", "휴대폰", "휴대전화", "핸드폰", "phone", "mobile", "tel", "contact"],
    "hostname": ["호스트", "호스트명", "hostname", "host", "서버명", "장비명"],
}
MAP_ORDER = ["ip", "asset_no", "dept", "owner", "contact", "hostname"]
BLANK_TOKENS = {"", "-", "--", ".", "n/a", "na", "없음", "미지정", "해당없음", "null"}


def clean_val(v) -> str:
    s = ("" if v is None else str(v)).strip()
    return "" if s.lower() in BLANK_TOKENS else s


def norm_header(s) -> str:
    return re.sub(r"[\s_\-./()]", "", ("" if s is None else str(s)).lower())


def unmerge_fill(ws) -> list[list]:
    """openpyxl 워크시트 → 병합해제된 AoA(병합 앵커값을 범위 전체로 채움)."""
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    width = max((len(r) for r in rows), default=0)
    for r in rows:
        r.extend([None] * (width - len(r)))
    for rng in list(ws.merged_cells.ranges):
        anchor = rows[rng.min_row - 1][rng.min_col - 1]
        for rr in range(rng.min_row - 1, rng.max_row):
            for cc in range(rng.min_col - 1, rng.max_col):
                rows[rr][cc] = anchor
    return rows


def detect_header_row(aoa) -> int:
    best, best_score = 0, -1
    lim = min(len(aoa), 25)
    for r in range(lim):
        vals = [str(c).strip() for c in (aoa[r] or []) if str("" if c is None else c).strip()]
        distinct = set(vals)
        if len(distinct) < 2:
            continue
        matched = 0
        for v in distinct:
            t = norm_header(v)
            for k in ASSET_ALIASES:
                if any(a in t for a in ASSET_ALIASES[k]):
                    matched += 1
                    break
        score = matched * 10 + len(distinct)
        if score > best_score:
            best_score, best = score, r
    return best


def columns_from(aoa, header_row):
    if not aoa:
        return []
    hr = max(0, min(header_row, len(aoa) - 1))
    header = aoa[hr]
    width = max((len(r) for r in aoa), default=0)
    data = [r for r in aoa[hr + 1:] if any(str("" if c is None else c).strip() for c in r)]
    cols = []
    for c in range(width):
        cols.append({
            "index": c,
            "header": (str(header[c]).strip() if c < len(header) and header[c] is not None else ""),
            "values": [(str(r[c]).strip() if c < len(r) and r[c] is not None else "") for r in data],
        })
    return cols


def auto_map(cols) -> dict:
    m = {}
    for col in cols:
        t = norm_header(col["header"])
        if not t:
            continue
        for k in MAP_ORDER:
            if k in m:
                continue
            if any(a in t for a in ASSET_ALIASES[k]):
                m[k] = col["index"]
                break
    return m


def resolve_cell(cols, spec, i) -> str:
    if spec is None:
        return ""
    if isinstance(spec, int):
        col, sep, part = spec, "", None
    else:
        col, sep, part = spec["col"], spec.get("sep", ""), spec.get("part")
    if col >= len(cols):
        return ""
    val = cols[col]["values"][i] if i < len(cols[col]["values"]) else ""
    if sep and part is not None:
        parts = str(val).split(sep)
        val = parts[part] if part < len(parts) else ""
    return clean_val(val)


KNOWN = ("asset_no", "hostname", "dept", "owner", "contact")


def build_records(cols, mapping) -> list[dict]:
    ip_spec = mapping.get("ip")
    if ip_spec is None:
        return []
    ip_col = ip_spec if isinstance(ip_spec, int) else ip_spec["col"]
    n = len(cols[ip_col]["values"]) if ip_col < len(cols) else 0
    out = []
    for i in range(n):
        ip = resolve_cell(cols, mapping.get("ip"), i)
        if not ip:
            continue
        rec = {"ip": ip, "asset_no": "", "dept": "", "owner": "", "contact": "", "hostname": "", "extra": {}}
        for k in KNOWN:
            if k in mapping:
                rec[k] = resolve_cell(cols, mapping[k], i)
        out.append(rec)
    return out


def process_sheet(ws, sheet_label, force_split_dept_owner=False):
    """워크시트 → (records, 매핑설명). 자산 데이터가 아니면 (None, 사유)."""
    aoa = unmerge_fill(ws)
    hr = detect_header_row(aoa)
    cols = columns_from(aoa, hr)
    mapping = auto_map(cols)
    if "ip" not in mapping:
        return None, f"IP 컬럼 없음(헤더행 {hr + 1}) — 자산 시트 아님으로 스킵"
    # 결합셀(소속 / 담당) 자동 분리 — dept 로 매핑된 컬럼 값이 대부분 '/' 를 포함하면
    # dept=part0, owner=part1 로 나눈다(프론트 매핑 UI 의 {col,sep,part} 와 동일).
    dept_spec = mapping.get("dept")
    if force_split_dept_owner and isinstance(dept_spec, int):
        vals = cols[dept_spec]["values"]
        if vals and sum("/" in v for v in vals) >= len(vals) * 0.5:
            mapping["dept"] = {"col": dept_spec, "sep": "/", "part": 0}
            mapping["owner"] = {"col": dept_spec, "sep": "/", "part": 1}
    recs = build_records(cols, mapping)
    desc = ", ".join(f"{k}→{('열' + str(v + 1) if isinstance(v, int) else '열' + str(v['col'] + 1) + '분리')}"
                     for k, v in mapping.items())
    return recs, f"헤더행 {hr + 1} · {desc} · 레코드 {len(recs)}건"


# ---------------------------------------------------------------------------
# HTTP
# ---------------------------------------------------------------------------
def req(base, method, path, token=None, body=None, xml_file=None):
    headers, data = {}, None
    if token:
        headers["Authorization"] = "Bearer " + token
    if body is not None:
        data = json.dumps(body, ensure_ascii=False).encode("utf-8")
        headers["Content-Type"] = "application/json; charset=utf-8"
    if xml_file is not None:
        boundary = "----scanopskit"
        data = (f"--{boundary}\r\n".encode()
                + f'Content-Disposition: form-data; name="file"; filename="{xml_file.name}"\r\n'.encode()
                + b"Content-Type: text/xml\r\n\r\n" + xml_file.read_bytes() + b"\r\n"
                + f"--{boundary}--\r\n".encode())
        headers["Content-Type"] = f"multipart/form-data; boundary={boundary}"
    r = urllib.request.Request(base + path, data=data, headers=headers, method=method)
    with urllib.request.urlopen(r, timeout=180) as resp:
        raw = resp.read().decode("utf-8")
        return json.loads(raw) if raw else None


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("password", help="admin 비밀번호 (data/INITIAL_ADMIN.txt)")
    ap.add_argument("--base", default="http://127.0.0.1:8770")
    args = ap.parse_args()
    base = args.base.rstrip("/")

    try:
        tok = req(base, "POST", "/api/auth/login",
                  body={"username": "admin", "password": args.password})["token"]
    except urllib.error.HTTPError as e:
        print(f"[!] 로그인 실패: {e.code} {e.read().decode('utf-8', 'replace')}")
        return 2
    print(f"[+] 로그인 OK ({base})")

    # 1) 스캔 XML 가져오기
    print("\n== 스캔 결과 가져오기 ==")
    for name, label in SCANS:
        path = HERE / name
        try:
            res = req(base, "POST", "/api/scans/import", tok, xml_file=path)
            print(f"  [OK]  {name:<28} {label:<20} counts={res.get('counts')}")
        except urllib.error.HTTPError as e:
            msg = e.read().decode("utf-8", "replace")
            tag = "예상된 거절" if name.startswith("scan_04") else "실패"
            print(f"  [{e.code}] {name:<28} {label:<20} {tag}: {msg[:80]}")

    # 2) dirty 자산대장 — 시트별 정제 후 업서트
    print("\n== dirty 자산대장 가져오기(고급 정제) ==")
    import openpyxl
    wb = openpyxl.load_workbook(HERE / "asset_ledger_dirty.xlsx", data_only=True)
    total_added = total_updated = total_matched = 0
    for sname in wb.sheetnames:
        ws = wb[sname]
        recs, desc = process_sheet(ws, sname, force_split_dept_owner=(sname == "인프라팀"))
        if recs is None:
            print(f"  [skip] 시트 '{sname}': {desc}")
            continue
        res = req(base, "POST", "/api/assets/bulk", tok, body=recs)
        total_added += res["added"]
        total_updated += res["updated"]
        print(f"  [OK]  시트 '{sname}': {desc}")
        print(f"        → 신규 {res['added']} · 갱신 {res['updated']} · 발견매칭 {res['findings_matched']}")

    # 3) 결과 요약
    print("\n== 적재 결과 요약 ==")
    findings = req(base, "GET", "/api/findings", tok)
    assets = req(base, "GET", "/api/assets", tok)
    from collections import Counter
    by_risk = Counter(f.get("risk_level", "?") for f in findings)
    by_dept = Counter((f.get("dept") or "(미매칭)") for f in findings)
    matched = sum(1 for f in findings if f.get("dept"))
    print(f"  자산 {len(assets)}건 · 발견 {len(findings)}건 · 자산 업서트 누계 신규 {total_added}·갱신 {total_updated}")
    print(f"  위험등급: {dict(by_risk)}")
    print(f"  부서매칭: 발견 {matched}/{len(findings)}건에 자산대장 부서 연결 · 상위 {dict(by_dept.most_common(6))}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
