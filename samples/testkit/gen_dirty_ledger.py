"""일부러 dirty 하게 작성된 자산대장 xlsx 생성기 (openpyxl 필요).

ScanOps 의 자산대장 고급 가져오기(`frontend/src/lib/assetImport.js` +
백엔드 `POST /assets/bulk`)가 현실의 지저분한 부서 취합본을 얼마나 견디는지
검증하기 위한 샘플. 아래 '더러움'을 의도적으로 심는다:

  · 제목/부제목 행이 헤더 위에 있음        → detectHeaderRow (헤더행 자동 감지)
  · 부서 셀 세로 병합(구간 반복)            → unmergeFillWs (병합해제 forward-fill)
  · 제목 셀 가로 병합                       → 병합 헤더 전파
  · 지저분하지만 별칭 매칭되는 헤더명        → computeAutoMap (자동 컬럼 매핑)
  · 누락/placeholder 토큰(-, N/A, 없음 …)   → cleanVal (빈값 정리)
  · 결합 셀(부서/담당 한 칸)                 → resolveCell {col,sep,part}
  · 중복 IP(업서트) · IP 누락 행(스킵)       → buildAssetRecords / bulk 업서트
  · 매핑 밖 여분 컬럼(도입일/위치/비고)      → extra(JSON) 보존
  · 앞뒤 공백 · 숫자형 셀 · 빈 행 섞임        → 정규화/trim

시트 3개:
  '자산대장(취합)'  메인 — 제목행+병합부서+지저분 헤더, 100행 이상
  '인프라팀'        부서 탭 — 결합셀(소속/담당), 헤더 순서 다름
  '요약'            자산 데이터가 아닌 피벗 표(잘못된 시트 선택 방지 시험)

사용: python gen_dirty_ledger.py   (출력: asset_ledger_dirty.xlsx)
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
OUT = HERE / "asset_ledger_dirty.xlsx"

BLANKS = ["-", "--", ".", "N/A", "na", "없음", "미지정", "해당없음", "null", ""]

# 스캔 샘플(scan_01_baseline)과 겹치는 IP — 가져오면 발견에 부서/담당/연락처 자동 매칭.
# (ip, hostname, dept, owner, contact, asset_no, note, location, intro_date)
SEED = [
    ("10.10.20.11", "sec-pc-01", "정보보안팀", "김보안", "02-1000-1101", "SEC-0001", "보안 관제 PC", "본관 3F 보안실", "2023-03-02"),
    ("10.10.20.12", "hr-srv-01", "인사팀", "박인사", "02-1000-1102", "HR-0007", "인사 포털 서버", "IDC A-12", "2022-11-20"),
    ("10.10.20.13", "infra-aix-01", "인프라팀", "이인프라", "010-2222-3333", "INF-0031", "AIX 기간계", "IDC B-03", "2019-06-15"),
    ("10.10.20.14", "fin-db-01", "재무팀", "최재무", "02-1000-1104", "FIN-0002", "회계 DB", "IDC A-05", "2021-01-11"),
    ("10.10.20.15", "web-proxy-01", "인프라팀", "이인프라", "010-2222-3333", "INF-0044", "리버스 프록시", "IDC B-07", "2024-02-01"),
    ("10.10.20.16", "misc-01", "개발팀", "정개발", "-", "DEV-0110", "테스트 노드", "본관 5F", "2024-09-09"),
    ("10.10.20.17", "iot-cam-01", "총무팀", "한총무", "없음", "GA-0301", "복도 CCTV", "본관 로비", "2020-05-05"),
    ("10.10.20.18", "dns-01", "인프라팀", "이인프라", "010-2222-3333", "INF-0009", "사내 DNS", "IDC B-01", "2018-08-08"),
    ("10.10.20.21", "", "정보보안팀", "김보안", "02-1000-1101", "SEC-0102", "PTR 없는 파일서버", "IDC A-20", "2023-07-19"),
    ("10.10.20.22", "legacy-01", "생산관리팀", "오생산", "N/A", "MFG-0007", "레거시 유닉스", "공장 1동", "2015-04-04"),
    ("10.10.20.23", "ops-vnc-01", "운영팀", "운영자", "미지정", "OPS-0055", "운영 콘솔(VNC)", "관제실", "2022-02-22"),
    ("10.10.20.26", "fw-shadow-01", "정보보안팀", "김보안", "02-1000-1101", "SEC-0203", "방화벽 그림자", "IDC A-01", "2023-12-30"),
]

DEPTS = ["정보보안팀", "인사팀", "재무팀", "인프라팀", "개발팀", "영업팀",
         "고객지원팀", "생산관리팀", "법무팀", "연구소", "총무팀", "운영팀"]
OWNERS = ["김철수", "이영희", "박민수", "최지훈", "정수빈", "강호민", "윤서연",
          "임재현", "한도윤", "오세훈", "서지우", "권나은"]


def _thin_msg():
    return "부서 취합 원본 — 셀 병합·빈칸 그대로 (정제는 가져오기에서)"


def build() -> Workbook:
    wb = Workbook()

    # ---- 시트 1: 메인 취합본 (dirty) ----
    ws = wb.active
    ws.title = "자산대장(취합)"

    # 컬럼: A 관리번호 | B IP 주소 | C 호스트명 | D 관리 부서 | E 담당자(성명)
    #       | F 연락처 / 내선 | G 도입일 | H 설치 위치 | I 비고
    NCOL = 9
    header = ["관리번호", "IP 주소", "호스트 명", "관리 부서", "담당자(성명)",
              "연락처 / 내선", "도입일", "설치 위치", "비 고"]

    # 1행: 가로 병합 제목
    ws.append(["2026년 상반기 정보자산 관리대장 (부서 취합본)"] + [None] * (NCOL - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOL)
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.cell(1, 1).alignment = Alignment(horizontal="center")
    # 2행: 부제(단일값에 가까운 행 — 헤더로 오인되면 안 됨)
    ws.append([_thin_msg()] + [None] * (NCOL - 1))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOL)
    ws.cell(2, 1).font = Font(italic=True, color="888888")
    # 3행: 진짜 헤더
    ws.append(header)
    for c in range(1, NCOL + 1):
        ws.cell(3, c).font = Font(bold=True)
        ws.cell(3, c).fill = PatternFill("solid", fgColor="DDE6F1")

    HEADER_ROW = 3
    rows: list[list] = []

    def add(asset_no, ip, host, dept, owner, contact, intro, loc, note):
        rows.append([asset_no, ip, host, dept, owner, contact, intro, loc, note])

    # (1) 스캔과 겹치는 시드 자산
    for ip, host, dept, owner, contact, ano, note, loc, intro in SEED:
        add(ano, ip, host, dept, owner, contact, intro, loc, note)

    # (2) 의도적 더러움 — 중복 IP(업서트), IP 누락(스킵), 앞뒤 공백, blank 토큰
    add("HR-0008", "  10.10.20.12  ", "hr-srv-01-dup", "인사팀", "박인사", "02-1000-1102",
        "2022-11-20", "IDC A-12", "★중복 IP(공백 포함) — 업서트 검증")
    add("DEV-0111", "", "no-ip-row", "개발팀", "정개발", "-", "2025-01-01", "본관 5F",
        "IP 없음 — 가져오기에서 스킵되어야")
    add("X-0000", "미지정", "blank-ip", "영업팀", "-", "없음", ".", "해당없음",
        "IP 자리에 placeholder — 스킵")
    add(9001, "10.10.20.31", "sales-pc-31", "영업팀", "  강영업  ", " 070-1234-5678 ",
        "2024-06-01", "지사 2F", "숫자형 관리번호 + 공백 값")

    # (3) 부서 세로 병합용 — 같은 부서 연속 블록(자동채움 검증). 100행 넘도록 대량 패딩.
    #     각 부서 블록의 첫 행만 부서명, 나머지는 빈칸 → 나중에 merge_cells 로 병합.
    blocks = [
        ("인프라팀", 14, "10.10.40."),
        ("개발팀", 12, "10.10.41."),
        ("재무팀", 10, "10.10.42."),
        ("고객지원팀", 11, "10.10.43."),
        ("생산관리팀", 13, "172.16.10."),
        ("연구소", 12, "172.16.20."),
        ("법무팀", 8, "172.16.30."),
        ("영업팀", 10, "10.10.44."),
    ]
    merge_dept_spans: list[tuple[int, int]] = []  # (start_row, end_row) 데이터행 기준(1-based sheet row)
    # 현재까지 rows 길이 → 다음 데이터가 들어갈 sheet row = HEADER_ROW + len(rows) + 1
    for dept, n, prefix in blocks:
        start_data_idx = len(rows)
        for k in range(n):
            octet = 11 + k
            ip = f"{prefix}{octet}"
            owner = OWNERS[(octet + len(dept)) % len(OWNERS)]
            # 일부 값은 blank 토큰/빈칸으로 더럽힘
            contact = BLANKS[(octet + k) % len(BLANKS)] if k % 4 == 0 else f"02-77{octet:02d}-{1000 + k}"
            note = BLANKS[k % len(BLANKS)] if k % 3 == 0 else f"{dept} 자산 #{k + 1}"
            host = "" if k % 5 == 0 else f"{dept[:3]}-srv-{octet}"
            intro = "" if k % 6 == 0 else f"202{k % 5}-0{1 + (k % 8)}-1{k % 9}"
            loc = BLANKS[(k * 2) % len(BLANKS)] if k % 7 == 0 else f"{dept} 구역 {k + 1}"
            ano = f"{dept[:2]}-{octet:04d}"
            # 부서명은 블록 첫 행에만(병합 앵커), 나머지는 빈칸
            dept_cell = dept if k == 0 else ""
            add(ano, ip, host, dept_cell, owner, contact, intro, loc, note)
        end_data_idx = len(rows) - 1
        s = HEADER_ROW + 1 + start_data_idx
        e = HEADER_ROW + 1 + end_data_idx
        if e > s:
            merge_dept_spans.append((s, e))

    # (4) 완전 빈 행 하나 섞기(빈 행 필터 검증)
    add("", "", "", "", "", "", "", "", "")
    add("ETC-0001", "10.10.20.99", "etc-99", "총무팀", "한총무", "02-1000-9999",
        "2026-07-01", "본관 1F", "말미 정상 행")

    # 데이터 행 적재
    for r in rows:
        ws.append(r)

    # 부서 세로 병합 적용(관리 부서 = D열 = 4)
    for s, e in merge_dept_spans:
        ws.merge_cells(start_row=s, start_column=4, end_row=e, end_column=4)
        ws.cell(s, 4).alignment = Alignment(vertical="center")

    # 폭 정리(가독)
    widths = [12, 18, 16, 12, 14, 16, 12, 16, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- 시트 2: 인프라팀 탭 (결합셀 + 헤더 순서 다름) ----
    ws2 = wb.create_sheet("인프라팀")
    ws2.append(["인프라팀 담당 자산 (별도 취합)"] + [None] * 5)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws2.append(["아이피", "자산코드", "장비명", "소속 / 담당", "phone", "note"])
    infra = [
        ("10.10.20.18", "INF-0009", "dns-01", "인프라팀 / 이인프라", "010-2222-3333", "사내 DNS(중복 시드 갱신)"),
        ("10.10.50.10", "INF-1001", "core-sw-01", "인프라팀 / 조네트", "02-7700-0010", "코어 스위치"),
        ("10.10.50.11", "INF-1002", "core-sw-02", "인프라팀 / 조네트", "02-7700-0011", "-"),
        ("10.10.50.12", "INF-1003", "san-01", "인프라팀 / 배스토리지", "없음", "SAN 스토리지"),
        ("10.10.50.13", "INF-1004", "backup-01", "인프라팀 / 배스토리지", ".", "백업 서버"),
        ("10.10.50.14", "INF-1005", "ntp-01", "인프라팀 / 이인프라", "N/A", "시각 동기"),
    ]
    for octet in range(20, 40):  # 인프라 패딩(20행+)
        infra.append((f"10.10.51.{octet}", f"INF-2{octet:03d}", f"esxi-{octet}",
                      f"인프라팀 / {OWNERS[octet % len(OWNERS)]}",
                      BLANKS[octet % len(BLANKS)] if octet % 3 == 0 else f"02-7712-{octet:04d}",
                      "가상화 호스트"))
    for row in infra:
        ws2.append(list(row))
    for i, w in enumerate([16, 12, 14, 22, 16, 24], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ---- 시트 3: 요약 (자산 데이터 아님 — 잘못된 시트) ----
    ws3 = wb.create_sheet("요약")
    ws3.append(["부서별 자산 집계 (요약)"])
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws3.append(["부서", "자산수", "비율"])
    ws3.append(["인프라팀", 42, "38%"])
    ws3.append(["개발팀", 18, "16%"])
    ws3.append(["재무팀", 12, "11%"])
    ws3.append(["합계", 112, "100%"])

    return wb


def main() -> None:
    wb = build()
    wb.save(OUT)
    # 데이터 행 수 계산(메인 시트 헤더 3행 제외)
    ws = wb["자산대장(취합)"]
    data_rows = ws.max_row - 3
    print(f"wrote {OUT.name} : 시트 {len(wb.sheetnames)}개 {wb.sheetnames}, "
          f"메인 데이터행 {data_rows}행 (헤더 3행)")


if __name__ == "__main__":
    main()
