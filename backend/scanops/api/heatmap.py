"""스캔 시간축 히트맵 — 저장된 nmap XML 로 phase/현재포트/4시트 보고서를 계산."""
from __future__ import annotations

import glob
import io
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..config import get_settings
from ..db import get_db
from ..models import Finding, RISK_LABELS_KO, ScanRun, User
from ..scanning.nmap_parse import scan_start
from ..spreadsheet import safe_cell
from .deps import current_user

router = APIRouter()
_settings = get_settings()

STATE_NEW_OPEN = "신규열림"
STATE_KEEP_OPEN = "기존열림"
STATE_NEW_CLOSED = "신규닫힘"
STATE_KEEP_CLOSED = "기존닫힘"
STATE_OUT_OF_SCOPE = "대상 외"

OPEN_STATES = {STATE_NEW_OPEN, STATE_KEEP_OPEN}
CLOSED_STATES = {STATE_NEW_CLOSED, STATE_KEEP_CLOSED}

STATE_COLORS = {
    STATE_NEW_OPEN: "FFE6D9D4",
    STATE_KEEP_OPEN: "FFDDEAF7",
    STATE_NEW_CLOSED: "FFEAD8EE",
    STATE_KEEP_CLOSED: "FFF7F7F7",
    STATE_OUT_OF_SCOPE: "FFEDEFF2",
}
HEADER_FILL = "FFE8ECF5"
HIGH_FILL = "FFFFE0D6"


def _split_key(key: str) -> tuple[str, str, int]:
    host, port, proto = key.split("|", 2)
    return host, proto, int(port)


def _display_label(scan: ScanRun) -> str:
    stamp = scan.started_at.strftime("%Y-%m-%d %H:%M") if scan.started_at else f"scan-{scan.id}"
    name = f" · {scan.name}" if scan.name else ""
    return f"#{scan.id} {stamp}{name}"


def _scan_xml_paths(scan: ScanRun) -> list[Path]:
    """ScanRun 이 남긴 XML 파일들. import/raw 는 scan_N.xml, chunk 는 scan_N.b*.xml."""
    seen: set[Path] = set()
    paths: list[Path] = []
    candidates: list[Path] = []
    if scan.raw_xml_path:
        candidates.append(Path(scan.raw_xml_path))
    base = _settings.scans_dir / f"scan_{scan.id}"
    candidates.append(Path(str(base) + ".xml"))
    candidates.extend(Path(p) for p in sorted(glob.glob(str(base) + ".b*.xml")))
    for p in candidates:
        try:
            rp = p.resolve()
        except OSError:
            continue
        if rp in seen or not p.exists():
            continue
        seen.add(rp)
        paths.append(p)
    return paths


def _text(svc, key: str) -> str:
    return (svc.get(key) if svc is not None else "") or ""


def _parse_xml_rows(path: Path) -> list[dict]:
    root = ET.parse(path).getroot()
    rows: list[dict] = []
    for host in root.findall("host"):
        status = host.find("status")
        if status is not None and status.get("state") == "down":
            continue
        addr_el = host.find("address[@addrtype='ipv4']")
        if addr_el is None:
            addr_el = host.find("address")
        host_ip = addr_el.get("addr") if addr_el is not None else ""
        if not host_ip:
            continue
        hn_el = host.find("hostnames/hostname")
        hostname = hn_el.get("name") if hn_el is not None else ""
        ports = host.find("ports")
        if ports is None:
            continue
        for port in ports.findall("port"):
            proto = (port.get("protocol") or "tcp").lower()
            try:
                port_num = int(port.get("portid") or "0")
            except ValueError:
                continue
            st = port.find("state")
            state = ((st.get("state") if st is not None else "open") or "open").lower()
            svc = port.find("service")
            product = _text(svc, "product")
            version = _text(svc, "version")
            extrainfo = _text(svc, "extrainfo")
            ostype = _text(svc, "ostype")
            detail = " ".join(x for x in (product, version, extrainfo, ostype) if x)
            service = _text(svc, "name")
            key = f"{host_ip}|{port_num}|{proto}"
            rows.append({
                "key": key,
                "host_ip": host_ip,
                "hostname": hostname,
                "proto": proto,
                "port": port_num,
                "state": state,
                "service": service,
                "product": product,
                "version": version,
                "banner": detail,
            })
    return rows


def _snapshots(db: Session) -> list[dict]:
    scans = (
        db.query(ScanRun)
        .filter(ScanRun.status == "done")
        .order_by(ScanRun.started_at, ScanRun.id)
        .all()
    )
    out: list[dict] = []
    for scan in scans:
        paths = _scan_xml_paths(scan)
        if not paths:
            continue
        rows_by_key: dict[str, dict] = {}
        scope_keys: set[str] = set()
        scope_ip_proto: set[tuple[str, str]] = set()
        scan_dates: list[datetime] = []
        for path in paths:
            try:
                if dt := scan_start(str(path)):
                    scan_dates.append(dt)
                for row in _parse_xml_rows(path):
                    rows_by_key[row["key"]] = row
                    scope_keys.add(row["key"])
                    scope_ip_proto.add((row["host_ip"], row["proto"]))
            except Exception:
                continue
        if not scope_keys:
            continue
        out.append({
            "scan": scan,
            "label": _display_label(scan),
            "scan_ids": [scan.id],
            "rows_by_key": rows_by_key,
            "scope_keys": scope_keys,
            "scope_ip_proto": scope_ip_proto,
            "open_keys": {k for k, r in rows_by_key.items() if r["state"].startswith("open")},
            "started_at": (scan_dates[0] if scan_dates else scan.started_at),
        })
    return out


def _group_phases(snapshots: list[dict]) -> list[dict]:
    phases: list[dict] = []
    cumulative_ip_proto: set[tuple[str, str]] = set()
    for snap in snapshots:
        scope_ip_proto = snap["scope_ip_proto"]
        if not phases or (scope_ip_proto & cumulative_ip_proto):
            phases.append({
                "labels": [snap["label"]],
                "scan_ids": list(snap["scan_ids"]),
                "rows_by_key": dict(snap["rows_by_key"]),
                "scope_keys": set(snap["scope_keys"]),
                "scope_ip_proto": set(scope_ip_proto),
                "open_keys": set(snap["open_keys"]),
                "started_at": snap["started_at"],
            })
        else:
            phase = phases[-1]
            phase["labels"].append(snap["label"])
            phase["scan_ids"].extend(snap["scan_ids"])
            phase["rows_by_key"].update(snap["rows_by_key"])
            phase["scope_keys"] |= snap["scope_keys"]
            phase["scope_ip_proto"] |= scope_ip_proto
            phase["open_keys"] |= snap["open_keys"]
        cumulative_ip_proto |= scope_ip_proto
    return phases


def _phase_label(phase: dict) -> str:
    labels = phase["labels"]
    if len(labels) == 1:
        return labels[0]
    return f"{labels[0]} ~ {labels[-1]} ({len(labels)}건)"


def _compute_states(phases: list[dict], keys: list[str]) -> tuple[dict[str, list[str]], dict[str, int | None]]:
    states: dict[str, list[str]] = {}
    last_idx: dict[str, int | None] = {}
    for key in keys:
        prev_open: bool | None = None
        tokens: list[str] = []
        last: int | None = None
        for idx, phase in enumerate(phases):
            if key in phase["scope_keys"]:
                last = idx
                if key in phase["open_keys"]:
                    tokens.append(STATE_KEEP_OPEN if prev_open is True else STATE_NEW_OPEN)
                    prev_open = True
                else:
                    tokens.append(STATE_NEW_CLOSED if prev_open is True else STATE_KEEP_CLOSED)
                    prev_open = False
            else:
                tokens.append(STATE_OUT_OF_SCOPE)
        states[key] = tokens
        last_idx[key] = last
    return states, last_idx


def _current_state(tokens: list[str]) -> str:
    for token in reversed(tokens):
        if token and token != STATE_OUT_OF_SCOPE:
            return token
    return ""


def _latest_row(phases: list[dict], key: str) -> dict:
    row: dict = {}
    for phase in phases:
        if key in phase["rows_by_key"]:
            row = phase["rows_by_key"][key]
    return row


def _last_open_row(phases: list[dict], key: str) -> dict:
    row: dict = {}
    for phase in phases:
        found = phase["rows_by_key"].get(key)
        if found and found["state"].startswith("open"):
            row = found
    return row


def _last_scan_label(phases: list[dict], key: str, last_idx: int | None) -> str:
    if last_idx is None:
        return ""
    return _phase_label(phases[last_idx])


def _key_sort(key: str):
    host, proto, port = _split_key(key)
    parts = []
    for p in host.split("."):
        try:
            parts.append(int(p))
        except ValueError:
            parts.append(p)
    return (parts, proto, port)


def build_heatmap(db: Session) -> dict:
    snapshots = _snapshots(db)
    phases = _group_phases(snapshots)
    finding_by_key = {f.finding_key: f for f in db.query(Finding).all()}
    ever_open = set()
    all_keys = set(finding_by_key)
    for snap in snapshots:
        all_keys |= set(snap["scope_keys"])
        ever_open |= set(snap["open_keys"])
    keys = sorted((ever_open | set(finding_by_key)), key=_key_sort)
    states, last_idx = _compute_states(phases, keys)

    rows = []
    for key in keys:
        host, proto, port = _split_key(key)
        finding = finding_by_key.get(key)
        latest = _latest_row(phases, key)
        last_open = _last_open_row(phases, key)
        token_list = states.get(key, [])
        current = _current_state(token_list)
        detail = last_open or latest
        risk = finding.risk_level if finding else ""
        service = (detail.get("service") or (finding.service if finding else "")) if detail else (finding.service if finding else "")
        version = (detail.get("version") or (finding.version if finding else "")) if detail else (finding.version if finding else "")
        rows.append({
            "key": key,
            "finding_id": finding.id if finding else None,
            "host_ip": host,
            "hostname": (finding.hostname if finding else "") or detail.get("hostname", ""),
            "proto": proto,
            "port": port,
            "service": service,
            "version": version,
            "risk_level": risk,
            "risk_label": RISK_LABELS_KO.get(risk, risk),
            "status": finding.status if finding else "",
            "dept": finding.dept if finding else "",
            "current_state": current,
            "observed_count": sum(1 for t in token_list if t != STATE_OUT_OF_SCOPE),
            "last_scan_label": _last_scan_label(phases, key, last_idx.get(key)),
            "cells": [{"state": token, "phase": i} for i, token in enumerate(token_list)],
        })

    current_open = [r for r in rows if r["current_state"] in OPEN_STATES]
    return {
        "states": {
            "new_open": STATE_NEW_OPEN,
            "keep_open": STATE_KEEP_OPEN,
            "new_closed": STATE_NEW_CLOSED,
            "keep_closed": STATE_KEEP_CLOSED,
            "out_of_scope": STATE_OUT_OF_SCOPE,
        },
        "phases": [
            {
                "index": i,
                "label": _phase_label(p),
                "scan_ids": p["scan_ids"],
                "scope_count": len(p["scope_keys"]),
                "open_count": len(p["open_keys"]),
            }
            for i, p in enumerate(phases)
        ],
        "rows": rows,
        "current_ports": current_open,
        "summary": {
            "scan_count": len(snapshots),
            "phase_count": len(phases),
            "row_count": len(rows),
            "current_open_count": len(current_open),
            "new_open_count": sum(1 for r in rows if r["current_state"] == STATE_NEW_OPEN),
            "new_closed_count": sum(1 for r in rows if r["current_state"] == STATE_NEW_CLOSED),
        },
    }


@router.get("")
def heatmap(_: User = Depends(current_user), db: Session = Depends(get_db)):
    return build_heatmap(db)


@router.get("/current")
def current_ports(_: User = Depends(current_user), db: Session = Depends(get_db)):
    h = build_heatmap(db)
    return {"total": len(h["current_ports"]), "items": h["current_ports"], "phases": h["phases"]}


def _append_sheet(wb, name: str, headers: list[str], rows: list[list], fills: list[list[str | None]] | None = None):
    from openpyxl.styles import PatternFill

    ws = wb.create_sheet(name)
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    for cell in ws[1]:
        cell.fill = header_fill
    for r_idx, row in enumerate(rows, start=2):
        ws.append([safe_cell(v) for v in row])
        if fills and r_idx - 2 < len(fills):
            for c_idx, color in enumerate(fills[r_idx - 2], start=1):
                if color:
                    ws.cell(r_idx, c_idx).fill = PatternFill("solid", fgColor=color)
    ws.freeze_panes = "A2"
    for col in ws.columns:
        width = min(max(len(str(c.value or "")) for c in col) + 2, 34)
        ws.column_dimensions[col[0].column_letter].width = width


def _report_bytes(data: dict) -> io.BytesIO:
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    summary = data["summary"]
    risk_counts: dict[str, int] = {}
    for row in data["current_ports"]:
        risk_counts[row["risk_label"] or row["risk_level"] or "미지정"] = risk_counts.get(row["risk_label"] or row["risk_level"] or "미지정", 0) + 1
    summary_rows = [
        ["스캔 수", summary["scan_count"]],
        ["phase 수", summary["phase_count"]],
        ["히트맵 행", summary["row_count"]],
        ["현재 열린 포트", summary["current_open_count"]],
        ["현재 신규열림", summary["new_open_count"]],
        ["현재 신규닫힘", summary["new_closed_count"]],
        [],
        ["위험등급", "현재 열린 포트 수"],
        *[[k, v] for k, v in sorted(risk_counts.items())],
    ]
    _append_sheet(wb, "00_보고요약", ["항목", "값"], summary_rows)

    phase_headers = [p["label"] for p in data["phases"]]
    heat_headers = ["IP", "프로토콜", "포트", "서비스", "위험도", "현재상태", "관측 phase 수", "마지막 스캔 시점"] + phase_headers
    heat_rows = []
    heat_fills = []
    for row in data["rows"]:
        tokens = [c["state"] for c in row["cells"]]
        heat_rows.append([
            row["host_ip"], row["proto"], row["port"], row["service"], row["risk_label"],
            row["current_state"], row["observed_count"], row["last_scan_label"], *tokens,
        ])
        fills = [None] * len(heat_headers)
        fills[5] = STATE_COLORS.get(row["current_state"])
        if row["risk_level"] in ("banned", "high"):
            fills[4] = HIGH_FILL
        for i, token in enumerate(tokens):
            fills[8 + i] = STATE_COLORS.get(token)
        heat_fills.append(fills)
    _append_sheet(wb, "01_시간축히트맵", heat_headers, heat_rows, heat_fills)

    current_headers = ["마지막 스캔 시점", "현재상태", "IP", "프로토콜", "포트", "서비스", "버전", "위험도", "운영상태", "부서"]
    current_rows = [
        [r["last_scan_label"], r["current_state"], r["host_ip"], r["proto"], r["port"], r["service"], r["version"], r["risk_label"], r["status"], r["dept"]]
        for r in data["current_ports"]
    ]
    current_fills = []
    for r in data["current_ports"]:
        fills = [None] * len(current_headers)
        fills[1] = STATE_COLORS.get(r["current_state"])
        if r["risk_level"] in ("banned", "high"):
            fills[7] = HIGH_FILL
        current_fills.append(fills)
    _append_sheet(wb, "02_현재포트현황", current_headers, current_rows, current_fills)

    cmp_headers = ["변경유형", "IP", "프로토콜", "포트", "서비스", "위험도", "첫 phase", "현재상태", "마지막 스캔 시점"]
    cmp_rows = []
    cmp_fills = []
    for r in data["rows"]:
        first = r["cells"][0]["state"] if r["cells"] else ""
        cur = r["current_state"]
        if first == cur and cur != STATE_NEW_OPEN:
            continue
        ctype = "신규열림" if cur in OPEN_STATES and first not in OPEN_STATES else "닫힘" if cur in CLOSED_STATES and first in OPEN_STATES else "상태변화"
        cmp_rows.append([ctype, r["host_ip"], r["proto"], r["port"], r["service"], r["risk_label"], first, cur, r["last_scan_label"]])
        fills = [STATE_COLORS.get(cur)] + [None] * (len(cmp_headers) - 1)
        cmp_fills.append(fills)
    _append_sheet(wb, "03_시점비교", cmp_headers, cmp_rows, cmp_fills)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/report")
def heatmap_report(_: User = Depends(current_user), db: Session = Depends(get_db)):
    buf = _report_bytes(build_heatmap(db))
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=scanops_heatmap_report.xlsx"},
    )
