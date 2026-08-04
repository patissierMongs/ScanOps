"""자산대장 라우터 — CRUD + xlsx 가져오기 + 발견 매칭(IP→부서)."""
from __future__ import annotations

import io
import zipfile

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy.orm import Session

from ..config import get_settings
from ..db import get_db
from ..models import Asset, Finding, User
from ..schemas import AssetIn, AssetOut
from ..uploads import read_limited
from .deps import current_user, require_role

router = APIRouter()
_settings = get_settings()

# xlsx 헤더 별칭 → 표준 필드
_ALIASES = {
    "ip": "ip", "아이피": "ip", "host_ip": "ip", "주소": "ip",
    "hostname": "hostname", "호스트명": "hostname", "호스트": "hostname",
    "dept": "dept", "부서": "dept", "담당부서": "dept",
    "owner": "owner", "담당자": "owner",
    "contact": "contact", "연락처": "contact", "전화": "contact",
    "asset_no": "asset_no", "자산번호": "asset_no", "관리번호": "asset_no",
    "note": "note", "비고": "note",
}

_ASSET_FIELDS = ("hostname", "dept", "owner", "contact", "asset_no", "note")


def _latest_assets_by_ip(db: Session) -> dict[str, Asset]:
    """Return the deterministic attribution source: newest asset row per IP."""
    by_ip: dict[str, Asset] = {}
    for asset in db.query(Asset).order_by(Asset.id).all():
        by_ip[asset.ip] = asset
    return by_ip


def _merge_extra(current: dict | None, patch: dict | None) -> dict:
    """Merge explicitly supplied extra keys; an explicit blank removes that key."""
    merged = dict(current or {})
    for raw_key, raw_value in (patch or {}).items():
        key = str(raw_key).strip()
        if not key:
            continue
        value = raw_value.strip() if isinstance(raw_value, str) else raw_value
        if value is None or value == "":
            merged.pop(key, None)
        else:
            merged[key] = value
    return merged


def _apply_asset_input(asset: Asset, body: AssetIn) -> None:
    """Apply only JSON fields the caller supplied, preserving omitted asset data."""
    supplied = body.model_fields_set
    for field in _ASSET_FIELDS:
        if field in supplied:
            value = getattr(body, field)
            setattr(asset, field, value.strip() if isinstance(value, str) else value)
    if "extra" in supplied:
        asset.extra = _merge_extra(asset.extra, body.extra)


def _validate_sheet_dimensions(rows: int, columns: int) -> None:
    if rows > _settings.asset_sheet_max_rows:
        raise HTTPException(
            status_code=413,
            detail=f"자산 시트가 허용 행 수({_settings.asset_sheet_max_rows})를 초과했습니다.",
        )
    if columns > _settings.asset_sheet_max_columns:
        raise HTTPException(
            status_code=413,
            detail=f"자산 시트가 허용 열 수({_settings.asset_sheet_max_columns})를 초과했습니다.",
        )
    if rows and columns and rows * columns > _settings.asset_sheet_max_cells:
        raise HTTPException(
            status_code=413,
            detail=f"자산 시트가 허용 셀 수({_settings.asset_sheet_max_cells})를 초과했습니다.",
        )


def _validate_xlsx_archive(data: bytes) -> None:
    """Reject malformed/oversized XLSX archives before openpyxl expands XML parts."""
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            entries = archive.infolist()
            if len(entries) > _settings.asset_xlsx_max_entries:
                raise HTTPException(
                    status_code=413,
                    detail=(
                        "자산 XLSX의 ZIP 항목 수가 허용 한도"
                        f"({_settings.asset_xlsx_max_entries}개)를 초과했습니다."
                    ),
                )
            expanded = 0
            for entry in entries:
                expanded += entry.file_size
                if expanded > _settings.asset_xlsx_max_uncompressed_bytes:
                    raise HTTPException(
                        status_code=413,
                        detail=(
                            "자산 XLSX의 압축 해제 크기가 허용 한도"
                            f"({_settings.asset_xlsx_max_uncompressed_bytes} bytes)를 초과했습니다."
                        ),
                    )
    except zipfile.BadZipFile as exc:
        raise HTTPException(status_code=400, detail="올바른 XLSX 파일이 아닙니다.") from exc


def match_assets(
    db: Session,
    ips: set[str] | None = None,
    *,
    clear_missing: bool = False,
    commit: bool = True,
) -> int:
    """Apply newest-asset attribution, optionally clearing IPs whose asset was removed.

    A finding department can also be assigned manually and is an operational field that must
    survive later scans.  Ordinary scan matching therefore leaves IPs with no asset untouched;
    asset deletion or an IP move opts into clearing stale asset-derived attribution.
    """
    by_ip = _latest_assets_by_ip(db)
    query = db.query(Finding)
    if ips is not None:
        if not ips:
            if commit:
                db.commit()
            else:
                db.flush()
            return 0
        query = query.filter(Finding.host_ip.in_(ips))
    n = 0
    for f in query.all():
        a = by_ip.get(f.host_ip)
        if a is None and not clear_missing:
            continue
        dept, contact, owner = (a.dept, a.contact, a.owner) if a else ("", "", "")
        changed = False
        if f.dept != dept:
            f.dept = dept
            changed = True
        if f.contact != contact:
            f.contact = contact
            changed = True
        if f.owner != owner:   # 자산대장 담당자명 → 통보에 활용
            f.owner = owner
            changed = True
        if changed:
            n += 1
    if commit:
        db.commit()
    else:
        db.flush()
    return n


@router.get("", response_model=list[AssetOut])
def list_assets(_: User = Depends(current_user), db: Session = Depends(get_db)):
    return db.query(Asset).order_by(Asset.ip).all()


@router.post("", response_model=AssetOut, status_code=201)
def create_asset(body: AssetIn, _: User = Depends(require_role("auditor")), db: Session = Depends(get_db)):
    ip = body.ip.strip()
    if not ip:
        raise HTTPException(status_code=400, detail="IP는 비울 수 없습니다.")
    a = Asset(ip=ip, extra={})
    _apply_asset_input(a, body)
    db.add(a)
    db.flush()
    match_assets(db, {ip})
    db.refresh(a)
    return a


@router.patch("/{aid}", response_model=AssetOut)
def update_asset(aid: int, body: AssetIn, _: User = Depends(require_role("auditor")), db: Session = Depends(get_db)):
    a = db.get(Asset, aid)
    if a is None:
        raise HTTPException(status_code=404, detail="자산을 찾을 수 없습니다.")
    ip = body.ip.strip()
    if not ip:
        raise HTTPException(status_code=400, detail="IP는 비울 수 없습니다.")
    old_ip = a.ip
    a.ip = ip
    _apply_asset_input(a, body)
    db.flush()
    match_assets(db, {old_ip, ip}, clear_missing=True)
    db.refresh(a)
    return a


@router.delete("/{aid}", status_code=204)
def delete_asset(aid: int, _: User = Depends(require_role("admin")), db: Session = Depends(get_db)):
    a = db.get(Asset, aid)
    if a is None:
        raise HTTPException(status_code=404, detail="자산을 찾을 수 없습니다.")
    old_ip = a.ip
    db.delete(a)
    db.flush()
    match_assets(db, {old_ip}, clear_missing=True)


@router.post("/bulk")
def bulk_import(
    body: list[AssetIn],
    _: User = Depends(require_role("auditor")),
    db: Session = Depends(get_db),
):
    """프론트(SheetJS)가 병합해제·헤더감지·매핑까지 끝낸 레코드를 IP 기준 업서트."""
    by_ip = _latest_assets_by_ip(db)
    added_ips: set[str] = set()
    updated_ips: set[str] = set()
    for rec in body:
        ip = rec.ip.strip()
        if not ip:
            continue
        a = by_ip.get(ip)
        if a is None:
            a = Asset(ip=ip, extra={})
            db.add(a)
            db.flush()
            by_ip[ip] = a
            added_ips.add(ip)
        elif ip not in added_ips:
            updated_ips.add(ip)
        _apply_asset_input(a, rec)
    db.flush()
    affected = added_ips | updated_ips
    matched = match_assets(db, affected)
    return {"added": len(added_ips), "updated": len(updated_ips), "findings_matched": matched}


@router.post("/import")
async def import_assets(
    file: UploadFile = File(...),
    _: User = Depends(require_role("auditor")),
    db: Session = Depends(get_db),
):
    import openpyxl

    data = await read_limited(file, _settings.upload_max_bytes)
    _validate_xlsx_archive(data)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="XLSX 통합문서를 읽을 수 없습니다.") from exc

    try:
        ws = wb.active
        declared_rows = int(ws.max_row or 0)
        declared_columns = int(ws.max_column or 0)
        _validate_sheet_dimensions(declared_rows, declared_columns)

        rows = iter(ws.iter_rows(values_only=True))
        try:
            header_row = next(rows)
        except StopIteration as exc:
            raise HTTPException(status_code=400, detail="빈 파일입니다.") from exc
        if not any(cell is not None and str(cell).strip() for cell in header_row):
            raise HTTPException(status_code=400, detail="빈 파일입니다.")
        _validate_sheet_dimensions(1, len(header_row))
        header = [(_ALIASES.get(str(c).strip().lower()) if c else None) for c in header_row]
        if "ip" not in header:
            raise HTTPException(status_code=400, detail="IP 컬럼을 찾을 수 없습니다. (헤더: IP/아이피/주소)")

        by_ip = _latest_assets_by_ip(db)
        added_ips: set[str] = set()
        updated_ips: set[str] = set()
        total_cells = len(header_row)
        for row_number, row in enumerate(rows, start=2):
            _validate_sheet_dimensions(row_number, len(row))
            total_cells += len(row)
            if total_cells > _settings.asset_sheet_max_cells:
                raise HTTPException(
                    status_code=413,
                    detail=f"자산 시트가 허용 셀 수({_settings.asset_sheet_max_cells})를 초과했습니다.",
                )
            values = {
                header[i]: (str(row[i]).strip() if row[i] is not None else "")
                for i in range(min(len(header), len(row)))
                if header[i]
            }
            ip = values.get("ip", "").strip()
            if not ip:
                continue
            rec = AssetIn(**{**values, "ip": ip})
            a = by_ip.get(ip)
            if a is None:
                a = Asset(ip=ip, extra={})
                db.add(a)
                db.flush()
                by_ip[ip] = a
                added_ips.add(ip)
            elif ip not in added_ips:
                updated_ips.add(ip)
            _apply_asset_input(a, rec)
        db.flush()
        affected = added_ips | updated_ips
        matched = match_assets(db, affected)
        return {"added": len(added_ips), "updated": len(updated_ips), "findings_matched": matched}
    finally:
        wb.close()
