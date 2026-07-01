"""Phase G 검증 — 대시보드 지표 + 감사 리포트(xlsx)."""
import io

import openpyxl

from tests.conftest import make_user, token_for

XML = "tests/fixtures/sample_scan.xml"


def _auth(client):
    make_user("op", "pw", role="auditor")
    return {"Authorization": f"Bearer {token_for(client, 'op', 'pw')}"}


def _import(client, h):
    with open(XML, "rb") as f:
        client.post("/api/scans/import", headers=h, files={"file": ("s.xml", f, "text/xml")})


def test_dashboard_metrics(client):
    h = _auth(client)
    _import(client, h)
    d = client.get("/api/dashboard", headers=h).json()
    assert d["open_total"] == 13
    assert d["by_risk"].get("high", 0) >= 1
    assert sum(d["by_status"].values()) == 13
    assert d["overdue"] == 0


def test_overdue_counts_after_deadline(client):
    h = _auth(client)
    _import(client, h)
    fid = client.get("/api/findings", headers=h).json()[0]["id"]
    # 과거 마감 → 초과로 집계
    client.patch(f"/api/findings/{fid}", headers=h,
                 json={"status": "처리중", "deadline": "2020-01-01T00:00:00"})
    d = client.get("/api/dashboard", headers=h).json()
    assert d["overdue"] == 1


def test_audit_report_xlsx(client):
    h = _auth(client)
    _import(client, h)
    r = client.get("/api/reports/audit", headers=h)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert ws.max_row == 14  # 헤더 + 13 발견
    assert ws.cell(1, 1).value == "발견키"
