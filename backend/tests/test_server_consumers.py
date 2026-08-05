"""Server 관측값이 사용자-facing 경로에서 일관되게 우선되는지 검증."""
from __future__ import annotations

import csv
import io

import openpyxl

from scanops.db import SessionLocal
from scanops.identity import display_identity
from scanops.models import Finding
from tests.conftest import make_user, token_for

XML = "tests/fixtures/sample_scan.xml"


def _auth(client):
    make_user("server-auditor", "pw", role="auditor")
    return {"Authorization": f"Bearer {token_for(client, 'server-auditor', 'pw')}"}


def _import_sample(client, headers):
    with open(XML, "rb") as source:
        response = client.post(
            "/api/scans/import", headers=headers,
            files={"file": ("sample.xml", source, "text/xml")},
        )
    assert response.status_code == 200, response.text


def _update_finding(port: int, **values):
    db = SessionLocal()
    try:
        row = db.query(Finding).filter(Finding.port == port).one()
        for key, value in values.items():
            setattr(row, key, value)
        db.commit()
    finally:
        db.close()


def _headers(ws) -> dict[str, int]:
    return {cell.value: cell.column for cell in ws[1]}


def test_display_identity_priority():
    assert display_identity(
        server="uvicorn", product="Generic HTTP", version="1.0", service="apple-iphoto",
    ) == "uvicorn"
    assert display_identity(product="OpenSSH", version="9.6", service="ssh") == "OpenSSH 9.6"
    assert display_identity(service="ssh") == "ssh"


def test_findings_search_matches_full_product_version_display_identity(client):
    headers = _auth(client)
    _import_sample(client, headers)
    _update_finding(
        9000, server="", product="OpenSSL", version="3.0", service="ssl",
    )

    response = client.get(
        "/api/findings", headers=headers, params={"q": "OpenSSL 3.0"},
    )

    assert response.status_code == 200, response.text
    assert [row["port"] for row in response.json()] == [9000]
    assert response.json()[0]["display_identity"] == "OpenSSL 3.0"

    csv_response = client.get(
        "/api/findings/export", headers=headers,
        params={"q": "OpenSSL 3.0", "fmt": "csv", "cols": "port,display_identity"},
    )
    assert csv_response.status_code == 200, csv_response.text
    csv_rows = list(csv.DictReader(io.StringIO(csv_response.content.decode("utf-8-sig"))))
    assert csv_rows == [{"포트": "9000", "표시 식별": "OpenSSL 3.0"}]

    xlsx_response = client.get(
        "/api/findings/export", headers=headers,
        params={"q": "OpenSSL 3.0", "fmt": "xlsx", "cols": "port,display_identity"},
    )
    assert xlsx_response.status_code == 200, xlsx_response.text
    ws = openpyxl.load_workbook(io.BytesIO(xlsx_response.content)).active
    assert list(ws.values) == [("포트", "표시 식별"), (9000, "OpenSSL 3.0")]


def test_global_event_feed_uses_display_identity_and_keeps_service_context(client):
    headers = _auth(client)
    _import_sample(client, headers)
    _update_finding(9000, server="uvicorn", service="apple-iphoto", product="", version="")
    finding = next(
        row for row in client.get("/api/findings", headers=headers).json()
        if row["port"] == 9000
    )
    changed = client.patch(
        f"/api/findings/{finding['id']}", headers=headers, json={"status": "처리중"},
    )
    assert changed.status_code == 200, changed.text

    feed = client.get("/api/events", headers=headers).json()
    item = next(row for row in feed["items"] if row["finding_id"] == finding["id"])

    assert item["display_identity"] == "uvicorn"
    assert item["server"] == "uvicorn"
    assert item["service"] == "apple-iphoto"


def test_default_csv_and_selected_xlsx_export_server_identity(client):
    headers = _auth(client)
    _import_sample(client, headers)
    _update_finding(9000, server="=1+1", service="http", product="Golang net/http server", version="")

    default_csv = client.get("/api/findings/export", headers=headers, params={"fmt": "csv"})
    assert default_csv.status_code == 200, default_csv.text
    csv_rows = list(csv.DictReader(io.StringIO(default_csv.content.decode("utf-8-sig"))))
    csv_row = next(row for row in csv_rows if row["포트"] == "9000")
    assert csv_row["표시 식별"] == "'=1+1"
    assert csv_row["Server"] == "'=1+1"
    assert csv_row["서비스"] == "http"

    selected_xlsx = client.get(
        "/api/findings/export", headers=headers,
        params={"fmt": "xlsx", "cols": "port,display_identity,server,service"},
    )
    assert selected_xlsx.status_code == 200, selected_xlsx.text
    ws = openpyxl.load_workbook(io.BytesIO(selected_xlsx.content)).active
    columns = _headers(ws)
    row_num = next(row for row in range(2, ws.max_row + 1) if ws.cell(row, columns["포트"]).value == 9000)
    assert ws.cell(row_num, columns["표시 식별"]).value == "'=1+1"
    assert ws.cell(row_num, columns["Server"]).value == "'=1+1"
    assert ws.cell(row_num, columns["서비스"]).value == "http"


def test_audit_report_includes_server_identity_and_escapes_formula(client):
    headers = _auth(client)
    _import_sample(client, headers)
    _update_finding(9000, server="=1+1", service="http", product="Golang net/http server", version="")

    response = client.get("/api/reports/audit", headers=headers)

    assert response.status_code == 200, response.text
    ws = openpyxl.load_workbook(io.BytesIO(response.content)).active
    columns = _headers(ws)
    row_num = next(row for row in range(2, ws.max_row + 1) if ws.cell(row, columns["포트"]).value == 9000)
    assert ws.cell(row_num, columns["표시 식별"]).value == "'=1+1"
    assert ws.cell(row_num, columns["Server"]).value == "'=1+1"
    assert ws.cell(row_num, columns["서비스"]).value == "http"


def test_notification_uses_server_then_labels_semantic_service(client):
    headers = _auth(client)
    client.post("/api/assets", headers=headers, json={"ip": "127.0.0.1", "dept": "인프라팀"})
    _import_sample(client, headers)
    _update_finding(9000, server="uvicorn", service="apple-iphoto", product="", version="")

    response = client.get(
        "/api/notifications/preview", headers=headers, params={"dept": "인프라팀"},
    )

    assert response.status_code == 200, response.text
    assert "127.0.0.1:9000/tcp uvicorn (서비스: apple-iphoto)" in response.json()["body"]


def test_heatmap_reuses_server_extractor_and_report_escapes_formula(client):
    headers = _auth(client)
    xml = b"""<?xml version="1.0"?>
<nmaprun start="1893456000">
  <host>
    <status state="up"/>
    <address addr="127.0.0.1" addrtype="ipv4"/>
    <ports>
      <port protocol="tcp" portid="8770">
        <state state="open"/>
        <service name="apple-iphoto" method="table" conf="3"/>
        <script id="http-server-header" output="uvicorn"/>
      </port>
      <port protocol="tcp" portid="8771">
        <state state="open"/>
        <service name="http" method="table" conf="3"/>
        <script id="http-server-header" output="=1+1"/>
      </port>
    </ports>
  </host>
</nmaprun>
"""
    imported = client.post(
        "/api/scans/import", headers=headers,
        files={"file": ("server.xml", xml, "text/xml")},
    )
    assert imported.status_code == 200, imported.text
    _update_finding(8770, server="")
    _update_finding(8771, server="")

    data = client.get("/api/heatmap", headers=headers).json()
    uvicorn = next(row for row in data["rows"] if row["port"] == 8770)
    assert uvicorn["server"] == "uvicorn"
    assert uvicorn["display_identity"] == "uvicorn"
    assert uvicorn["service"] == "apple-iphoto"

    report = client.get("/api/heatmap/report", headers=headers)
    assert report.status_code == 200, report.text
    ws = openpyxl.load_workbook(io.BytesIO(report.content))["02_현재포트현황"]
    columns = _headers(ws)
    row_num = next(row for row in range(2, ws.max_row + 1) if ws.cell(row, columns["포트"]).value == 8771)
    assert ws.cell(row_num, columns["표시 식별"]).value == "'=1+1"
    assert ws.cell(row_num, columns["Server"]).value == "'=1+1"
    assert ws.cell(row_num, columns["서비스"]).value == "http"


def test_findings_search_covers_ip_port_and_every_displayed_field(client):
    """발견 관리 검색은 '화면에 보이는 모든 요소'로 찾을 수 있어야 한다.

    한때 이 폭넓은 검색이 있었는데 Server 우선표시 작업 때 6개 컬럼으로 좁아졌다.
    IP·포트로 못 찾는 검색은 운영에서 사실상 쓸 수 없다."""
    headers = _auth(client)
    _import_sample(client, headers)
    _update_finding(
        9000, host_ip="10.9.9.9", hostname="wiki.corp", service="http",
        product="nginx", version="1.24", server="nginx/1.24",
        banner="Zimbra-Collab", cpe="cpe:/a:nginx:nginx", category="웹",
        usage="사내 위키", remarks="NSE: TLS_CN=wiki.corp",
        dept="정보보안팀", owner="김보안", contact="010-0000-0000",
        manual_note="차기 점검 대상",
    )

    def found(term):
        response = client.get("/api/findings", headers=headers, params={"q": term})
        assert response.status_code == 200, response.text
        return [row["port"] for row in response.json()]

    # IP(부분일치 포함)와 포트 — 사용자가 가장 자주 쓰는 두 축
    assert found("10.9.9.9") == [9000]
    assert found("10.9.9") == [9000]
    assert found("9000") == [9000]
    # 나머지 표시 필드 전부. 'http' 처럼 다른 발견과도 겹치는 값이 있으므로 포함 여부로 본다.
    for term in ("wiki.corp", "http", "nginx", "1.24", "nginx/1.24", "Zimbra-Collab",
                 "cpe:/a:nginx", "웹", "사내 위키", "TLS_CN", "정보보안팀", "김보안",
                 "010-0000-0000", "차기 점검"):
        assert 9000 in found(term), f"'{term}' 로 검색되지 않음"
    # 결합된 표시 식별자(product + version)도 유지
    assert found("nginx 1.24") == [9000]
    # 없는 값은 안 잡혀야 한다(전부 매칭되는 헛검색 방지)
    assert found("존재하지않는값") == []


def test_findings_search_ignores_surrounding_whitespace(client):
    headers = _auth(client)
    _import_sample(client, headers)
    _update_finding(9000, host_ip="10.9.9.9")

    response = client.get("/api/findings", headers=headers, params={"q": "  10.9.9.9  "})

    assert response.status_code == 200, response.text
    assert [row["port"] for row in response.json()] == [9000]


def test_findings_search_rejects_non_ascii_digits_without_crashing(client):
    """isdigit() 은 '②' 같은 유니코드 숫자에도 True 라 int() 에서 500 이 난다.
    isdecimal() 이어야 포트 변환이 안전하다."""
    headers = _auth(client)
    _import_sample(client, headers)

    response = client.get("/api/findings", headers=headers, params={"q": "９０００"})

    assert response.status_code == 200, response.text


def test_findings_export_honors_the_same_broad_search(client):
    headers = _auth(client)
    _import_sample(client, headers)
    _update_finding(9000, host_ip="10.9.9.9", dept="정보보안팀")

    response = client.get(
        "/api/findings/export", headers=headers,
        params={"q": "10.9.9.9", "fmt": "csv", "cols": "host_ip,port"},
    )

    assert response.status_code == 200, response.text
    rows = list(csv.DictReader(io.StringIO(response.content.decode("utf-8-sig"))))
    assert rows == [{"IP": "10.9.9.9", "포트": "9000"}]


def _lookup():
    return {
        "http": {"category": "웹", "usage": "웹 서비스", "risk_level": "medium",
                 "compliance": [{"std": "KISA", "ref": "웹"}]},
        "https": {"category": "웹", "usage": "웹 서비스(TLS)", "risk_level": "low",
                  "compliance": []},
        "ssh": {"category": "원격", "usage": "원격접속", "risk_level": "high", "compliance": []},
    }


def test_server_banner_classifies_findings_nmap_mislabels():
    """nmap 이 uniconv 처럼 저신뢰 추측을 내놔도 Server 헤더가 나왔다면 HTTP 로 분류한다.

    Server 헤더는 http-server-header/http-headers 가 실제 HTTP 응답을 받아냈다는 뜻이라
    service 추측보다 강한 증거다. taxonomy 는 제품명이 아니라 서비스명으로 키가 잡혀 있어
    '이 포트는 HTTP 로 말한다'는 사실만 키로 되돌린다."""
    from scanops.scanning.taxonomy import classify

    finding = {"service": "uniconv", "port": 8080, "server": "nginx/1.24.0",
               "nse_json": {"http-server-header": "nginx/1.24.0"}}

    classify(finding, _lookup(), [])

    assert finding["category"] == "웹"
    assert finding["usage"] == "웹 서비스"
    assert finding["risk_level"] == "medium"      # info 로 방치되지 않는다
    evidence = [c for c in finding["compliance_json"] if c["std"] == "관측근거"]
    assert evidence and "uniconv" in evidence[0]["ref"] and "nginx/1.24.0" in evidence[0]["ref"]


def test_server_banner_classification_uses_tls_evidence_for_https():
    from scanops.scanning.taxonomy import classify

    for nse in ({"ssl-cert": "commonName=x"}, '{"ssl-cert": "commonName=x"}'):
        finding = {"service": "apple-iphoto", "port": 8443, "server": "nginx", "nse_json": nse}
        classify(finding, _lookup(), [])
        assert finding["usage"] == "웹 서비스(TLS)", nse


def test_server_banner_never_overrides_a_service_that_already_classifies():
    """보조 키일 뿐이라 기존에 잘 분류되던 건의 위험등급을 흔들지 않는다."""
    from scanops.scanning.taxonomy import classify

    finding = {"service": "ssh", "port": 22, "server": "nginx", "nse_json": {}}

    classify(finding, _lookup(), [])

    assert finding["category"] == "원격" and finding["risk_level"] == "high"
    assert not [c for c in finding["compliance_json"] if c["std"] == "관측근거"]


def test_unclassifiable_finding_without_server_stays_untouched():
    from scanops.scanning.taxonomy import classify

    finding = {"service": "uniconv", "port": 9999, "server": "", "nse_json": {}}

    classify(finding, _lookup(), [])

    assert finding["category"] == "" and finding["risk_level"] == "info"


def test_reclassify_all_applies_server_fallback(client):
    """재계산 경로에도 관측 증거가 전달돼야 한다(예전엔 service/port 만 넘겼다)."""
    from scanops.scanning.taxonomy import reclassify_all

    headers = _auth(client)
    _import_sample(client, headers)
    _update_finding(9000, service="uniconv", server="nginx/1.24.0",
                    nse_json={"http-server-header": "nginx/1.24.0"},
                    category="", usage="", risk_level="info")

    db = SessionLocal()
    try:
        reclassify_all(db)
        row = db.query(Finding).filter(Finding.port == 9000).one()
        assert row.category == "웹"
        assert row.risk_level != "info"
    finally:
        db.close()
