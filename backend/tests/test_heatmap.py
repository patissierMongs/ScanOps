"""스캔 시간축 히트맵 — XML 기반 상태 계산과 XLSX 보고서."""
from __future__ import annotations

import io
from datetime import timezone
from pathlib import Path

import openpyxl
from scanops.api import scans as scans_api
from scanops.db import SessionLocal
from scanops.models import Finding, ScanRun
from scanops.scanning.nmap_parse import scan_start
from tests.conftest import make_user, token_for

SAMPLES = Path(__file__).resolve().parents[2] / "samples"


def _auth(client):
    make_user("op", "pw", role="auditor")
    return {"Authorization": f"Bearer {token_for(client, 'op', 'pw')}"}


def _import(client, headers, path: Path, name: str | None = None):
    with path.open("rb") as f:
        return client.post(
            "/api/scans/import",
            headers=headers,
            files={"file": (name or path.name, f, "text/xml")},
        )


def _row(data: dict, port: int) -> dict:
    return next(r for r in data["rows"] if r["host_ip"] == "127.0.0.1" and r["port"] == port)


def test_heatmap_tracks_open_and_closed_ports(client):
    headers = _auth(client)
    assert _import(client, headers, SAMPLES / "scanA.xml").status_code == 200
    assert _import(client, headers, SAMPLES / "scanB.xml").status_code == 200

    data = client.get("/api/heatmap", headers=headers).json()

    assert data["summary"]["scan_count"] == 2
    assert data["summary"]["phase_count"] == 2

    port_3000 = _row(data, 3000)
    assert [c["state"] for c in port_3000["cells"]] == ["신규열림", "신규닫힘"]
    assert port_3000["current_state"] == "신규닫힘"

    port_8080 = _row(data, 8080)
    assert [c["state"] for c in port_8080["cells"]] == ["신규열림", "기존열림"]
    assert port_8080["current_state"] == "기존열림"


def test_narrow_port_scan_does_not_overwrite_heatmap_current(client, tmp_path):
    headers = _auth(client)
    assert _import(client, headers, SAMPLES / "scanA.xml").status_code == 200
    narrow_xml = tmp_path / "narrow.xml"
    narrow_xml.write_text(
        """<?xml version="1.0"?>
<nmaprun start="1893456000">
  <host>
    <status state="up"/>
    <address addr="127.0.0.1" addrtype="ipv4"/>
    <ports>
      <port protocol="tcp" portid="1">
        <state state="closed"/>
        <service name="tcpmux" method="table" conf="3"/>
      </port>
    </ports>
  </host>
</nmaprun>
""",
        encoding="utf-8",
    )
    assert _import(client, headers, narrow_xml, "narrow.xml").status_code == 200

    data = client.get("/api/heatmap", headers=headers).json()
    port_3000 = _row(data, 3000)

    assert [c["state"] for c in port_3000["cells"]] == ["신규열림", "대상 외"]
    assert port_3000["current_state"] == "신규열림"

    current = client.get("/api/heatmap/current", headers=headers).json()
    assert any(r["host_ip"] == "127.0.0.1" and r["port"] == 3000 for r in current["items"])


def test_staged_heatmap_uses_sweep_fallback_then_selected_rescan_force_close(
    client, monkeypatch, tmp_path,
):
    monkeypatch.setattr(scans_api._settings, "data_dir", tmp_path)
    scans_api._settings.ensure_dirs()
    headers = _auth(client)
    key = "127.0.0.1|18443|tcp"
    sweep_xml = (
        '<?xml version="1.0"?><nmaprun><host><status state="up"/>'
        '<address addr="127.0.0.1" addrtype="ipv4"/><ports>'
        '<port protocol="tcp" portid="18443"><state state="open"/>'
        '<service name="unknown" method="table"/></port>'
        '</ports></host></nmaprun>'
    )

    db = SessionLocal()
    try:
        existing = Finding(
            finding_key=key, host_ip="127.0.0.1", port=18443, proto="tcp", state="open",
            service="http", product="Uvicorn", version="0.30", server="uvicorn 0.30",
            nse_json=[{"id": "http-server-header", "output": "uvicorn 0.30"}],
        )
        opened = ScanRun(name="staged open", status="running")
        db.add_all([existing, opened])
        db.commit()
        open_dir = scans_api._settings.scans_dir / f"scan_{opened.id}"
        open_dir.mkdir(parents=True, exist_ok=True)
        (open_dir / "stage-tcp-b0.xml").write_text(sweep_xml, encoding="utf-8")

        scans_api._commit_engine_ingest(db, opened, open_dir, {key}, False)
        opened.status = "done"
        db.commit()
        opened_id = opened.id
        expected_start = opened.started_at.replace(tzinfo=timezone.utc)
        snapshot_path = Path(opened.raw_xml_path)
    finally:
        db.close()

    assert int(scan_start(snapshot_path).timestamp()) == int(expected_start.timestamp())
    first = client.get("/api/heatmap", headers=headers).json()
    first_row = _row(first, 18443)
    assert first["phases"][0]["scan_ids"] == [opened_id]
    assert first_row["current_state"] == "신규열림"
    assert first_row["display_identity"] == "uvicorn 0.30"
    assert first_row["server"] == "uvicorn 0.30"
    assert client.get("/api/heatmap/current", headers=headers).json()["total"] == 1

    db = SessionLocal()
    try:
        unreachable = ScanRun(name="ordinary scan unreachable", status="running")
        db.add(unreachable)
        db.commit()
        unreachable_dir = scans_api._settings.scans_dir / f"scan_{unreachable.id}"
        unreachable_dir.mkdir(parents=True, exist_ok=True)

        scans_api._commit_engine_ingest(db, unreachable, unreachable_dir, {key}, False)
        unreachable.status = "done"
        db.commit()
    finally:
        db.close()

    closed_by_scope = client.get("/api/heatmap", headers=headers).json()
    closed_by_scope_row = _row(closed_by_scope, 18443)
    # A completed structured scan owns its explicit target/port scope even when discovery
    # observes no live host, so its snapshot records the requested finding as closed.
    assert [cell["state"] for cell in closed_by_scope_row["cells"]] == ["신규열림", "신규닫힘"]
    assert closed_by_scope_row["current_state"] == "신규닫힘"
    assert client.get("/api/heatmap/current", headers=headers).json()["total"] == 0

    db = SessionLocal()
    try:
        closed = ScanRun(name="selected rescan close", status="running")
        db.add(closed)
        db.commit()
        close_dir = scans_api._settings.scans_dir / f"scan_{closed.id}"
        close_dir.mkdir(parents=True, exist_ok=True)
        # A selected rescan must ignore an unrelated/stale full-scan sweep artifact. With no
        # active stage3 row, the selected scope key is authoritatively closed.
        (close_dir / "stage-tcp-b0.xml").write_text(sweep_xml, encoding="utf-8")

        scans_api._commit_engine_ingest(db, closed, close_dir, {key}, True)
        closed.status = "done"
        db.commit()
    finally:
        db.close()

    final = client.get("/api/heatmap", headers=headers).json()
    final_row = _row(final, 18443)
    assert [cell["state"] for cell in final_row["cells"]] == ["신규열림", "신규닫힘", "기존닫힘"]
    assert final_row["current_state"] == "기존닫힘"
    assert client.get("/api/heatmap/current", headers=headers).json()["total"] == 0


def test_heatmap_report_xlsx_has_operational_sheets(client):
    headers = _auth(client)
    assert _import(client, headers, SAMPLES / "scanA.xml").status_code == 200
    assert _import(client, headers, SAMPLES / "scanB.xml").status_code == 200

    res = client.get("/api/heatmap/report", headers=headers)

    assert res.status_code == 200
    assert "spreadsheetml" in res.headers["content-type"]
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    assert wb.sheetnames == ["00_보고요약", "01_시간축히트맵", "02_현재포트현황", "03_시점비교"]
    assert wb["01_시간축히트맵"].cell(1, 1).value == "IP"


def test_heatmap_key_sort_mixes_ip_hostname_and_ipv6_without_crashing():
    """호스트명 대상이 하나라도 있으면 히트맵 전체가 500 이 되던 회귀 방지.

    예전 정렬키는 옥텟을 '되는 것만' int 로 바꿔 [10,0,0,5] 와 ['web01'] 같은 혼합 타입
    리스트를 만들었고, sorted() 가 int 와 str 을 비교하다 TypeError 를 냈다."""
    from scanops.api.heatmap import _key_sort

    keys = [
        "10.0.0.10|22|tcp", "web01.local|443|tcp", "10.0.0.2|22|tcp",
        "fe80::1|22|tcp", "10.0.0.2|80|tcp", "srv|22|tcp",
    ]

    ordered = sorted(keys, key=_key_sort)   # TypeError 나면 여기서 실패

    hosts = [key.split("|")[0] for key in ordered]
    # IPv4 는 사전순이 아니라 수치순(10.0.0.2 가 10.0.0.10 보다 앞)
    assert hosts.index("10.0.0.2") < hosts.index("10.0.0.10")
    # IP 가 이름 기반 대상보다 앞
    assert hosts.index("10.0.0.10") < hosts.index("web01.local")
    assert set(hosts) == {"10.0.0.10", "web01.local", "10.0.0.2", "fe80::1", "srv"}


def test_heatmap_endpoint_survives_hostname_targets(client):
    """정렬 크래시가 API 레벨에서도 재현되지 않는지(사용자가 본 GET /api/heatmap 500)."""
    from scanops.db import SessionLocal
    from scanops.models import Finding

    headers = _auth(client)
    assert _import(client, headers, SAMPLES / "scanA.xml").status_code == 200
    db = SessionLocal()
    try:
        db.add(Finding(
            finding_key="web01.local|443|tcp", host_ip="web01.local", port=443,
            proto="tcp", state="open", service="https", status="열림",
        ))
        db.commit()
    finally:
        db.close()

    response = client.get("/api/heatmap", headers=headers)

    assert response.status_code == 200, response.text
    assert any(row["host_ip"] == "web01.local" for row in response.json()["rows"])
